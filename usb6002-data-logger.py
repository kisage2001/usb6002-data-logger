"""
USB-6002 Data Acquisition System
High-performance data logger with real-time monitoring

Developer: Tsutsumi Hiroki
Institution: Tokyo National College of Technology
Version: 2.4
Date: 2026-05-29

Changes from v2.3:
- [Change] Unified Digital I/O panel. P0.0–P0.7 share the same physical lines,
           so each line now has a per-line mode toggle (row 1): IN (cyan) or
           OUT (orange). Row 2 shows state: for OUT lines, click toggles output
           yellow(OFF)->green(ON); for IN lines, the line is read continuously
           at the plot draw cadence and shows yellow(LOW)/green(HIGH). DO and DI
           tasks are rebuilt dynamically from the mode assignment, so a line is
           never used as input and output simultaneously.

Changes from v2.2:
- [Change] Digital Input is now always-on: clicking a line number arms it and
           it is read automatically at the plot draw cadence (~60 ms) inside the
           existing AI timer loop — no separate timer, no Start/Stop button.
           HIGH=green, LOW=red. DI reads are skipped during Bode measurement
           (update_plot is paused then), so no interference with AI sampling.
- [Safety] DO and DI share the same 8 physical lines; a line armed for DI is
           excluded from the DO task and vice-versa, preventing contention.

Changes from v2.1:
- [New] Digital Input panel: monitor P0.0–P0.7 as inputs. Click a line number
        to arm it, then Start DI Monitor to poll (100 ms) and show HIGH (green)
        / LOW (gray) state by color. Placed directly below the Digital Output
        panel.
- [UI]  Controls reorganized: Record/Hold/Clear on row 1, Clipboard/Save/Export
        on row 2, Settings centered on row 3. DO All ON/OFF moved into the DO
        panel. Left control panel is scrollable.

Changes from v2.0:
- [New] Frequency Response (Bode) measurement: AO0 step-sweep excitation,
        selectable input-reference (AO0 command or any AI channel, e.g. a motor
        driver command signal) and response (any AI channel). Per-frequency
        single-sine least-squares fit yields gain (dB) and phase (deg).
        Configurable f start/stop, points, amplitude, settle time, measurement
        cycles, and log/linear sweep. Results saved to Excel (data + Bode image)
        and shown in a separate Bode window with adjustable graph format
        (title, font family, font sizes, line width, marker size, grid).

Changes from v1.9:
- [New] Digital Output panel: P0.0–P0.7 toggle buttons (compact single row).
- [UI]  Control buttons reorganized into a clean 3x3 grid.

Changes from v1.5:
- [New] Function Generator: AO0/AO1 の2chファンクションジェネレータ機能を追加。
        左パネル下部に2ch並列レイアウトで配置。
        各chで波形（Sine/Square/Triangle/Sawtooth/DC）、周波数、振幅、
        オフセット、位相を独立して設定可能。
        出力ボタン押下で色が変わり出力開始。AIモニタリングと同時動作可能。
        AO更新レート 5 kS/s（USB-6002 仕様上限）、実用周波数上限 ~500 Hz。

Changes from v1.4:
- [New] Hold Snapshot: pressing Hold now automatically saves the currently
        displayed buffer data to an Excel file (Hold_Snapshot_YYYYMMDD_HHMMSS.xlsx)
        in the same data directory (C:\\PRG\\data).

Changes from v1.1:
- [Fix] Active channel states (checkboxes) are now saved to config.json and
        restored on next launch.
- [Fix] Bulk-change in ConfigEditor now covers Unit, Y Min, and Y Max in
        addition to Conversion Factor.
- [Fix] Hold/Resume: the DAQ buffer is now drained during Hold so that
        resuming does not flood the plot with accumulated samples.
"""

import nidaqmx
from nidaqmx.constants import TerminalConfiguration, AcquisitionType, READ_ALL_AVAILABLE
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import matplotlib
matplotlib.use('TkAgg')
import pandas as pd
import numpy as np
import datetime
import os
import collections
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
from PIL import Image
import io

# バージョン情報
__version__ = "2.4"
__author__ = "Tsutsumi Hiroki"
__date__ = "2026-05-29"
__institution__ = "Tokyo National College of Technology"

# デフォルト設定ファイルのパス
CONFIG_FILE = "config.json"

# カラーパレット（論文用：色覚多様性対応）
COLORS = ['#E41A1C', '#377EB8', '#4DAF4A', '#984EA3',
          '#FF7F00', '#A65628', '#F781BF', '#999999']


class ConfigEditor(tk.Toplevel):
    """設定エディタウィンドウ"""
    def __init__(self, parent, config):
        super().__init__(parent)
        self.title("Configuration Editor")
        self.geometry("800x650")
        self.config = config
        self.result = None

        # メインフレーム
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ノートブック（タブ）
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # チャンネル設定タブ
        channel_frame = ttk.Frame(notebook, padding="10")
        notebook.add(channel_frame, text="Channels")

        # グラフ設定タブ
        graph_frame = ttk.Frame(notebook, padding="10")
        notebook.add(graph_frame, text="Graph Settings")

        # === チャンネル設定タブの内容 ===
        canvas = tk.Canvas(channel_frame)
        scrollbar = ttk.Scrollbar(channel_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # ------------------------------------------------------------------ #
        # row=0: カラムヘッダー
        # row=1: 一括変更入力（各列の真下に配置）
        # row=2〜: チャンネルごとの設定値
        # ------------------------------------------------------------------ #

        # --- row 0: ヘッダー ---
        COL_CH     = 0
        COL_NAME   = 1
        COL_FACTOR = 2
        COL_UNIT   = 3
        COL_YMIN   = 4
        COL_YMAX   = 5

        WIDTHS = {
            COL_CH:     4,
            COL_NAME:   20,
            COL_FACTOR: 15,
            COL_UNIT:   10,
            COL_YMIN:   10,
            COL_YMAX:   10,
        }

        headers = {
            COL_CH:     "CH",
            COL_NAME:   "Name",
            COL_FACTOR: "Conv. Factor",
            COL_UNIT:   "Unit",
            COL_YMIN:   "Y Min",
            COL_YMAX:   "Y Max",
        }
        for col, text in headers.items():
            ttk.Label(scrollable_frame, text=text,
                      font=('Arial', 10, 'bold')).grid(
                row=0, column=col, padx=5, pady=(8, 2), sticky=tk.W)

        # --- row 1: 一括変更行 ---
        # CH列: ラベル
        ttk.Label(scrollable_frame, text="ALL",
                  font=('Arial', 9, 'italic'),
                  foreground='gray').grid(
            row=1, column=COL_CH, padx=5, pady=2)

        # Name列: 一括変更なし（空白）

        # Conv. Factor
        self.bulk_factor_var = tk.StringVar(value="")
        bulk_factor_frame = ttk.Frame(scrollable_frame)
        bulk_factor_frame.grid(row=1, column=COL_FACTOR, padx=5, pady=2, sticky=tk.W)
        ttk.Entry(bulk_factor_frame, textvariable=self.bulk_factor_var,
                  width=WIDTHS[COL_FACTOR] - 5).pack(side=tk.LEFT)
        ttk.Button(bulk_factor_frame, text="▶", width=2,
                   command=self.apply_bulk_factor).pack(side=tk.LEFT, padx=(2, 0))

        # Unit
        self.bulk_unit_var = tk.StringVar(value="")
        bulk_unit_frame = ttk.Frame(scrollable_frame)
        bulk_unit_frame.grid(row=1, column=COL_UNIT, padx=5, pady=2, sticky=tk.W)
        ttk.Entry(bulk_unit_frame, textvariable=self.bulk_unit_var,
                  width=WIDTHS[COL_UNIT] - 3).pack(side=tk.LEFT)
        ttk.Button(bulk_unit_frame, text="▶", width=2,
                   command=self.apply_bulk_unit).pack(side=tk.LEFT, padx=(2, 0))

        # Y Min
        self.bulk_ymin_var = tk.StringVar(value="")
        bulk_ymin_frame = ttk.Frame(scrollable_frame)
        bulk_ymin_frame.grid(row=1, column=COL_YMIN, padx=5, pady=2, sticky=tk.W)
        ttk.Entry(bulk_ymin_frame, textvariable=self.bulk_ymin_var,
                  width=WIDTHS[COL_YMIN] - 3).pack(side=tk.LEFT)
        ttk.Button(bulk_ymin_frame, text="▶", width=2,
                   command=self.apply_bulk_ymin).pack(side=tk.LEFT, padx=(2, 0))

        # Y Max
        self.bulk_ymax_var = tk.StringVar(value="")
        bulk_ymax_frame = ttk.Frame(scrollable_frame)
        bulk_ymax_frame.grid(row=1, column=COL_YMAX, padx=5, pady=2, sticky=tk.W)
        ttk.Entry(bulk_ymax_frame, textvariable=self.bulk_ymax_var,
                  width=WIDTHS[COL_YMAX] - 3).pack(side=tk.LEFT)
        ttk.Button(bulk_ymax_frame, text="▶", width=2,
                   command=self.apply_bulk_ymax).pack(side=tk.LEFT, padx=(2, 0))

        # 区切り線
        ttk.Separator(scrollable_frame, orient='horizontal').grid(
            row=2, column=0, columnspan=6, sticky=tk.EW, padx=5, pady=4)

        # チャンネル設定の入力フィールド
        self.channel_entries = []
        for i, ch in enumerate(config['channels']):
            row = i + 3  # header(0) + bulk(1) + separator(2)

            ttk.Label(scrollable_frame, text=f"CH{ch['id']}").grid(
                row=row, column=0, padx=5, pady=2)

            name_var = tk.StringVar(value=ch['name'])
            ttk.Entry(scrollable_frame, textvariable=name_var,
                      width=20).grid(row=row, column=1, padx=5, pady=2)

            factor_var = tk.DoubleVar(value=ch['conversion_factor'])
            ttk.Entry(scrollable_frame, textvariable=factor_var,
                      width=15).grid(row=row, column=2, padx=5, pady=2)

            unit_var = tk.StringVar(value=ch['unit'])
            ttk.Entry(scrollable_frame, textvariable=unit_var,
                      width=10).grid(row=row, column=3, padx=5, pady=2)

            ymin_var = tk.DoubleVar(value=ch['y_min'])
            ttk.Entry(scrollable_frame, textvariable=ymin_var,
                      width=10).grid(row=row, column=4, padx=5, pady=2)

            ymax_var = tk.DoubleVar(value=ch['y_max'])
            ttk.Entry(scrollable_frame, textvariable=ymax_var,
                      width=10).grid(row=row, column=5, padx=5, pady=2)

            self.channel_entries.append({
                'id': ch['id'],
                'name': name_var,
                'factor': factor_var,
                'unit': unit_var,
                'ymin': ymin_var,
                'ymax': ymax_var
            })

        # === グラフ設定タブの内容 ===
        font_frame = ttk.LabelFrame(graph_frame, text="Font Settings", padding="10")
        font_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=10, pady=10)

        ttk.Label(font_frame, text="Font Family:").grid(
            row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.font_family_var = tk.StringVar(
            value=config['graph'].get('font_family', 'Arial'))
        font_family_combo = ttk.Combobox(
            font_frame, textvariable=self.font_family_var, width=20)
        font_family_combo['values'] = (
            'Arial', 'Times New Roman', 'Helvetica', 'Courier New',
            'Verdana', 'Georgia', 'Comic Sans MS', 'DejaVu Sans',
            'Liberation Sans', 'Noto Sans')
        font_family_combo.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(font_frame, text="Title Font Size:").grid(
            row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.title_fontsize_var = tk.IntVar(
            value=config['graph']['font_size_title'])
        ttk.Spinbox(font_frame, from_=8, to=32,
                    textvariable=self.title_fontsize_var,
                    width=10).grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(font_frame, text="Axis Label Font Size:").grid(
            row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.label_fontsize_var = tk.IntVar(
            value=config['graph']['font_size_label'])
        ttk.Spinbox(font_frame, from_=8, to=24,
                    textvariable=self.label_fontsize_var,
                    width=10).grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(font_frame, text="Tick Label Font Size:").grid(
            row=3, column=0, sticky=tk.W, padx=5, pady=5)
        self.tick_fontsize_var = tk.IntVar(
            value=config['graph']['font_size_tick'])
        ttk.Spinbox(font_frame, from_=6, to=20,
                    textvariable=self.tick_fontsize_var,
                    width=10).grid(row=3, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(font_frame, text="Legend Font Size:").grid(
            row=4, column=0, sticky=tk.W, padx=5, pady=5)
        self.legend_fontsize_var = tk.IntVar(
            value=config['graph']['font_size_legend'])
        ttk.Spinbox(font_frame, from_=6, to=20,
                    textvariable=self.legend_fontsize_var,
                    width=10).grid(row=4, column=1, sticky=tk.W, padx=5, pady=5)

        other_frame = ttk.LabelFrame(graph_frame, text="Other Settings", padding="10")
        other_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=10, pady=10)

        ttk.Label(other_frame, text="Graph Title:").grid(
            row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.graph_title_var = tk.StringVar(value=config['graph']['title'])
        ttk.Entry(other_frame, textvariable=self.graph_title_var,
                  width=40).grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(other_frame, text="X-axis Label:").grid(
            row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.xlabel_var = tk.StringVar(value=config['graph']['xlabel'])
        ttk.Entry(other_frame, textvariable=self.xlabel_var,
                  width=40).grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(other_frame, text="Y-axis Label (Left):").grid(
            row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.ylabel_var = tk.StringVar(
            value=config['graph'].get('ylabel', 'Voltage (V)'))
        ttk.Entry(other_frame, textvariable=self.ylabel_var,
                  width=40).grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(other_frame, text="Y-axis Label (Right):").grid(
            row=3, column=0, sticky=tk.W, padx=5, pady=5)
        self.ylabel_right_var = tk.StringVar(
            value=config['graph'].get('ylabel_right', ''))
        ttk.Entry(other_frame, textvariable=self.ylabel_right_var,
                  width=40).grid(row=3, column=1, padx=5, pady=5)
        ttk.Label(other_frame, text="(empty = hidden)",
                  font=('Arial', 8), foreground='gray').grid(
            row=3, column=2, sticky=tk.W, padx=2, pady=5)

        ttk.Label(other_frame, text="Show Grid:").grid(
            row=4, column=0, sticky=tk.W, padx=5, pady=5)
        self.grid_var = tk.BooleanVar(value=config['graph']['grid'])
        ttk.Checkbutton(other_frame, variable=self.grid_var).grid(
            row=4, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(other_frame, text="Line Width:").grid(
            row=5, column=0, sticky=tk.W, padx=5, pady=5)
        self.linewidth_var = tk.DoubleVar(value=config['graph']['line_width'])
        ttk.Spinbox(other_frame, from_=0.5, to=5.0, increment=0.5,
                    textvariable=self.linewidth_var,
                    width=10).grid(row=5, column=1, sticky=tk.W, padx=5, pady=5)

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(button_frame, text="OK",
                   command=self.ok_clicked).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="Cancel",
                   command=self.cancel_clicked).pack(side=tk.RIGHT)

    # ---------------------------------------------------------------------- #
    # 一括変更ヘルパー
    # ---------------------------------------------------------------------- #
    def apply_bulk_factor(self):
        """変換係数を全チャンネルに一括適用"""
        try:
            value = float(self.bulk_factor_var.get())
            for entry in self.channel_entries:
                entry['factor'].set(value)
            messagebox.showinfo("Success",
                                f"Conversion factor {value} applied to all channels.")
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number.")

    def apply_bulk_unit(self):
        """単位を全チャンネルに一括適用"""
        value = self.bulk_unit_var.get().strip()
        if not value:
            messagebox.showerror("Error", "Please enter a unit string.")
            return
        for entry in self.channel_entries:
            entry['unit'].set(value)
        messagebox.showinfo("Success",
                            f"Unit '{value}' applied to all channels.")

    def apply_bulk_ymin(self):
        """Y Min を全チャンネルに一括適用"""
        try:
            value = float(self.bulk_ymin_var.get())
            for entry in self.channel_entries:
                entry['ymin'].set(value)
            messagebox.showinfo("Success",
                                f"Y Min {value} applied to all channels.")
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number.")

    def apply_bulk_ymax(self):
        """Y Max を全チャンネルに一括適用"""
        try:
            value = float(self.bulk_ymax_var.get())
            for entry in self.channel_entries:
                entry['ymax'].set(value)
            messagebox.showinfo("Success",
                                f"Y Max {value} applied to all channels.")
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number.")

    # ---------------------------------------------------------------------- #
    def ok_clicked(self):
        try:
            for i, entry in enumerate(self.channel_entries):
                self.config['channels'][i]['name'] = entry['name'].get()
                self.config['channels'][i]['conversion_factor'] = entry['factor'].get()
                self.config['channels'][i]['unit'] = entry['unit'].get()
                self.config['channels'][i]['y_min'] = entry['ymin'].get()
                self.config['channels'][i]['y_max'] = entry['ymax'].get()

            self.config['graph']['font_family'] = self.font_family_var.get()
            self.config['graph']['font_size_title'] = self.title_fontsize_var.get()
            self.config['graph']['font_size_label'] = self.label_fontsize_var.get()
            self.config['graph']['font_size_tick'] = self.tick_fontsize_var.get()
            self.config['graph']['font_size_legend'] = self.legend_fontsize_var.get()
            self.config['graph']['title'] = self.graph_title_var.get()
            self.config['graph']['xlabel'] = self.xlabel_var.get()
            self.config['graph']['ylabel'] = self.ylabel_var.get()
            self.config['graph']['ylabel_right'] = self.ylabel_right_var.get()
            self.config['graph']['grid'] = self.grid_var.get()
            self.config['graph']['line_width'] = self.linewidth_var.get()

            self.result = self.config
            self.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Invalid input:\n{e}")

    def cancel_clicked(self):
        self.result = None
        self.destroy()


# ========================================================================== #
class DAQApplication:
    def __init__(self, root):
        self.root = root
        self.root.title(f"USB-6002 Data Acquisition System v{__version__}")
        self.root.geometry("1400x900")

        # 設定の読み込み
        self.config = self.load_config()

        # 初期化
        self.task = None
        self.is_recording = False
        self.is_paused = False
        self.recorded_data = None
        self.recording_count = 0
        self.target_samples = 0
        self.update_id = None
        self.legend = None

        self.update_counter = 0
        self.draw_interval = 3   # 描画頻度を抑えて軽量化（20ms×3≒15fps）
        self._time_axis_cache = None
        self._time_axis_cache_len = -1

        self.channel_name_vars = []
        self.channel_factor_vars = []

        # ------------------------------------------------------------------ #
        # [Fix 1] active_channels を config から復元
        # ------------------------------------------------------------------ #
        saved_active = self.config.get('active_channels', [True] * 8)
        # チャンネル数が変わっていても対応できるよう長さを揃える
        self.active_channels = (saved_active + [True] * 8)[:8]

        self.current_values = [0.0] * 8
        self.offsets = [0.0] * 8
        self.offset_labels = [None] * 8

        # アクティブchインデックスキャッシュ（update_plotで毎回再計算しない）
        self._active_ch_indices = [
            i for i, a in enumerate(self.active_channels) if a]

        # ------------------------------------------------------------------ #
        # [New v1.6] Function Generator: AOタスク管理
        # ------------------------------------------------------------------ #
        self.ao_tasks   = [None, None]     # AO0, AO1 それぞれのnidaqmx.Task
        self.ao_running = [False, False]   # 出力中フラグ
        self.ao_out_btns = [None, None]    # 出力ボタン参照（色変更用）
        self.ao_preview_after = [None, None]  # デバウンス用after ID

        # ------------------------------------------------------------------ #
        # [New v2.4] Digital I/O 統合管理（P0.0〜P0.7）
        #   各ラインは 'in' または 'out' モード。
        #   OUTタスク・INタスクをモードに応じて動的に構築する。
        # ------------------------------------------------------------------ #
        self.dio_mode = ['out'] * 8        # 各ラインのモード 'in' / 'out'
        self.dio_out_state = [False] * 8   # OUTライン: False=OFF(黄) True=ON(緑)
        self.dio_in_state = [False] * 8    # INライン: 直近の読取り値
        self.dio_mode_btns = [None] * 8    # モードボタン参照
        self.dio_state_btns = [None] * 8   # 状態ボタン参照

        self.do_task = None                # 出力タスク
        self._do_lines = []                # 出力タスクが保持するライン
        self.di_task = None                # 入力タスク
        self._di_lines = []                # 入力タスクが保持するライン
        self._dio_dirty = True             # タスク再構築フラグ

        # ------------------------------------------------------------------ #
        # [New v2.1] Frequency Response 測定
        # ------------------------------------------------------------------ #
        self.fr_running = False            # 測定中フラグ
        self.fr_results = None             # 直近の測定結果（dict）
        self.fr_graph_cfg = {              # ボード線図フォーマット（変更可能）
            'font_family': 'Arial',
            'font_size_title': 14,
            'font_size_label': 12,
            'font_size_tick': 10,
            'title': 'Bode Diagram',
            'line_width': 1.5,
            'marker_size': 5,
            'grid': True,
        }

        # ------------------------------------------------------------------ #
        # [New v2.1] Frequency Response (Bode) 測定
        # ------------------------------------------------------------------ #
        self.freq_resp_running = False     # 測定実行中フラグ
        self.freq_resp_data = None         # 測定結果 (freqs, gains, phases)
        self.freq_resp_btn = None          # 開始/停止ボタン参照

        self.create_widgets()
        self.setup_graph()
        self.start_task()
        self.update_plot()

    # ---------------------------------------------------------------------- #
    # 設定読み込み
    # ---------------------------------------------------------------------- #
    def load_config(self):
        default_config = {
            "device": {
                "device_name": "Dev1",
                "sampling_rate": 1000,
                "samples_per_channel": 100,
                "terminal_config": "RSE"
            },
            "channels": [
                {"id": i, "name": f"CH{i}", "conversion_factor": 1.0,
                 "unit": "V", "y_min": -10, "y_max": 10}
                for i in range(8)
            ],
            # [Fix 1] active_channels をデフォルト設定に追加
            "active_channels": [True] * 8,
            "graph": {
                "title": "Real-time Monitoring",
                "xlabel": "Time [s]",
                "ylabel": "Voltage (V)",
                "ylabel_right": "",
                "time_window": 5.0,
                "grid": True,
                "line_width": 1.5,
                "font_family": "Arial",
                "font_size_title": 18,
                "font_size_label": 16,
                "font_size_tick": 14,
                "font_size_legend": 12,
                "legend_position": "upper right",
                "legend_columns": 2
            }
        }

        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                for key in default_config:
                    if key not in config:
                        config[key] = default_config[key]
                    elif isinstance(default_config[key], dict):
                        for subkey in default_config[key]:
                            if subkey not in config[key]:
                                config[key][subkey] = default_config[key][subkey]
                return config
            except Exception as e:
                print(f"Warning: Failed to load config file: {e}")
                return default_config
        return default_config

    # ---------------------------------------------------------------------- #
    # UI 構築
    # ---------------------------------------------------------------------- #
    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="6")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ── 左パネルをスクロール可能にする（画面が小さくても全機能にアクセス可）── #
        left_container = ttk.Frame(main_frame, width=440)
        left_container.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        left_container.pack_propagate(False)

        left_canvas = tk.Canvas(left_container, highlightthickness=0, width=440)
        left_scroll = ttk.Scrollbar(left_container, orient="vertical",
                                    command=left_canvas.yview)
        control_frame = ttk.Frame(left_canvas)
        control_frame.bind(
            "<Configure>",
            lambda e: left_canvas.configure(scrollregion=left_canvas.bbox("all")))
        left_canvas.create_window((0, 0), window=control_frame, anchor="nw", width=424)
        left_canvas.configure(yscrollcommand=left_scroll.set)
        left_canvas.pack(side="left", fill="both", expand=True)
        left_scroll.pack(side="right", fill="y")

        # マウスホイールで左パネルをスクロール
        def _on_left_wheel(event):
            left_canvas.yview_scroll(int(-event.delta / 120), "units")
        left_canvas.bind("<Enter>",
                         lambda e: left_canvas.bind_all("<MouseWheel>", _on_left_wheel))
        left_canvas.bind("<Leave>",
                         lambda e: left_canvas.unbind_all("<MouseWheel>"))

        # ── ヘッダー（最小化：1行のみ）──────────────────────────────────── #
        header_frame = ttk.Frame(control_frame)
        header_frame.pack(fill=tk.X, pady=(0, 2))
        ttk.Label(header_frame, text=f"USB-6002 DAQ  v{__version__}",
                  font=('Arial', 10, 'bold')).pack(anchor=tk.W)

        # ── ステータス ────────────────────────────────────────────────────── #
        status_frame = ttk.LabelFrame(control_frame, text="Status", padding="2")
        status_frame.pack(fill=tk.X, pady=(0, 3))
        self.status_label = ttk.Label(status_frame, text="Monitoring...",
                                      font=('Arial', 10, 'bold'), foreground='green')
        self.status_label.pack()
        self.progress_label = ttk.Label(status_frame, text="", font=('Arial', 8))
        self.progress_label.pack()

        # ── ボタン類を整理（操作系をまとめ、余白を確保）──────────────────── #
        btn_frame = ttk.LabelFrame(control_frame, text="Controls", padding="3")
        btn_frame.pack(fill=tk.X, pady=(0, 3))
        for c in range(3):
            btn_frame.columnconfigure(c, weight=1, uniform="ctrl")

        # 1行目：測定操作
        self.record_btn = ttk.Button(btn_frame, text="🔴 Record",
                                     command=self.start_recording)
        self.record_btn.grid(row=0, column=0, sticky=tk.EW, padx=1, pady=1)
        self.hold_btn = ttk.Button(btn_frame, text="⏸ Hold",
                                   command=self.toggle_hold)
        self.hold_btn.grid(row=0, column=1, sticky=tk.EW, padx=1, pady=1)
        ttk.Button(btn_frame, text="🔄 Clear",
                   command=self.clear_all_data).grid(
            row=0, column=2, sticky=tk.EW, padx=1, pady=1)

        # 2行目：出力・設定
        ttk.Button(btn_frame, text="📋 Clipboard",
                   command=self.copy_to_clipboard).grid(
            row=1, column=0, sticky=tk.EW, padx=1, pady=1)
        ttk.Button(btn_frame, text="💾 Save Fig",
                   command=self.save_figure).grid(
            row=1, column=1, sticky=tk.EW, padx=1, pady=1)
        ttk.Button(btn_frame, text="📊 Export",
                   command=self.quick_export_menu).grid(
            row=1, column=2, sticky=tk.EW, padx=1, pady=1)

        # 3行目：設定（単独・中央寄せ）
        ttk.Button(btn_frame, text="⚙ Settings",
                   command=self.open_settings).grid(
            row=2, column=1, sticky=tk.EW, padx=1, pady=1)

        # ── Digital I/O（P0.0〜P0.7）統合パネル ──────────────────────────── #
        #   1行目：各ラインのモード IN/OUT をトグル（IN=水色, OUT=橙）
        #   2行目：状態 — OUT は押すとトグル出力(黄→緑)、IN は読取り(OFF黄/ON緑)
        dio_outer = ttk.LabelFrame(control_frame,
                                   text="Digital I/O (P0.0–P0.7)", padding="3")
        dio_outer.pack(fill=tk.X, pady=(0, 3))

        # モード行
        mode_grid = ttk.Frame(dio_outer)
        mode_grid.pack(fill=tk.X)
        ttk.Label(mode_grid, text="Mode", font=('Arial', 7, 'bold'),
                  width=5).grid(row=0, column=0, sticky=tk.W)
        for c in range(8):
            mode_grid.columnconfigure(c + 1, weight=1, uniform="dio")
        for line in range(8):
            mb = tk.Button(mode_grid, text="OUT",
                           font=('Arial', 7, 'bold'),
                           width=3, padx=0, pady=0, bd=1,
                           bg='#e8923a', fg='white', relief=tk.RAISED,
                           command=lambda ln=line: self.toggle_dio_mode(ln))
            mb.grid(row=0, column=line + 1, padx=1, pady=1, sticky=tk.EW)
            self.dio_mode_btns[line] = mb

        # 状態行
        state_grid = ttk.Frame(dio_outer)
        state_grid.pack(fill=tk.X)
        ttk.Label(state_grid, text="State", font=('Arial', 7, 'bold'),
                  width=5).grid(row=0, column=0, sticky=tk.W)
        for c in range(8):
            state_grid.columnconfigure(c + 1, weight=1, uniform="dio")
        for line in range(8):
            sb = tk.Button(state_grid, text=f"{line}",
                           font=('Arial', 9, 'bold'),
                           width=3, padx=0, pady=0, bd=1,
                           bg='#f0d000', fg='black', relief=tk.RAISED,
                           command=lambda ln=line: self.on_dio_state_click(ln))
            sb.grid(row=0, column=line + 1, padx=1, pady=1, sticky=tk.EW)
            self.dio_state_btns[line] = sb

        ttk.Label(dio_outer,
                  text="IN(cyan)/OUT(orange) | state: yellow=OFF/LOW  green=ON/HIGH",
                  font=('Arial', 7), foreground='gray').pack(anchor=tk.W, pady=(1, 0))

        # ── Active Channels（内部スクロールなし：左パネル全体がスクロール）── #
        channel_frame = ttk.LabelFrame(control_frame, text="Active Channels",
                                       padding="3")
        channel_frame.pack(fill=tk.X, pady=(0, 3))

        scrollable_frame = ttk.Frame(channel_frame)
        scrollable_frame.pack(fill=tk.X)

        hdr = ttk.Frame(scrollable_frame)
        hdr.pack(fill=tk.X, pady=(0, 3))
        for col, (txt, w) in enumerate(
                [("On", 3), ("Name", 8), ("Factor", 7), ("Current", 11),
                 ("Zero/Reset", 9), ("Offset", 14)]):
            ttk.Label(hdr, text=txt, font=('Arial', 9, 'bold'),
                      width=w).grid(row=0, column=col, padx=1)

        self.channel_vars = []
        self.channel_name_vars = []
        self.channel_factor_vars = []
        self.value_labels = []
        self.offset_labels = []    # オフセット値表示ラベル
        self.zero_reset_btns = []  # Zero/Reset トグルボタン

        for i, ch in enumerate(self.config['channels']):
            # --- row 0: チェック / 名前 / 係数 / 現在値 ---
            ch_row = ttk.Frame(scrollable_frame)
            ch_row.pack(fill=tk.X, pady=(2, 0))

            # [Fix 1] 保存済みのアクティブ状態をチェックボックスの初期値に使用
            var = tk.BooleanVar(value=self.active_channels[i])
            ttk.Checkbutton(ch_row, variable=var,
                            command=self.update_active_channels,
                            width=3).grid(row=0, column=0, padx=1)
            self.channel_vars.append(var)

            name_var = tk.StringVar(value=ch['name'])
            self.channel_name_vars.append(name_var)
            name_entry = ttk.Entry(ch_row, textvariable=name_var, width=8)
            name_entry.grid(row=0, column=1, padx=1)
            name_entry.bind('<Return>', lambda e, idx=i: self.update_channel_name(idx))
            name_entry.bind('<FocusOut>', lambda e, idx=i: self.update_channel_name(idx))

            factor_var = tk.DoubleVar(value=ch['conversion_factor'])
            self.channel_factor_vars.append(factor_var)
            factor_entry = ttk.Entry(ch_row, textvariable=factor_var, width=7)
            factor_entry.grid(row=0, column=2, padx=1)
            factor_entry.bind('<Return>',
                              lambda e, idx=i: self.update_channel_factor(idx))
            factor_entry.bind('<FocusOut>',
                              lambda e, idx=i: self.update_channel_factor(idx))

            value_frame = ttk.Frame(ch_row)
            value_frame.grid(row=0, column=3, padx=1)
            tk.Canvas(value_frame, width=10, height=10,
                      highlightthickness=0, bg=COLORS[i]).pack(
                side=tk.LEFT, padx=(0, 3))
            val_label = ttk.Label(value_frame,
                                  text=f"+0.000 {ch['unit']}",
                                  font=('Courier', 9), foreground=COLORS[i],
                                  width=11)
            val_label.pack(side=tk.LEFT)
            self.value_labels.append(val_label)

            # Zero/Reset トグルボタン（col=4）
            # オフセット未設定："Zero"（通常）、設定済み："Reset"（赤）
            zr_btn = ttk.Button(ch_row, text="Zero", width=9,
                                command=lambda idx=i: self.toggle_offset(idx))
            zr_btn.grid(row=0, column=4, padx=(1, 1))
            self.zero_reset_btns.append(zr_btn)

            # オフセット値ラベル（col=5）
            off_label = ttk.Label(ch_row,
                                  text="---",
                                  font=('Courier', 8), foreground='gray',
                                  width=14)
            off_label.grid(row=0, column=5, padx=1, sticky=tk.W)
            self.offset_labels.append(off_label)

        # ================================================================== #
        # [New v1.6] Function Generator パネル（AO0 / AO1 横並び）
        # [Update v1.7] 波形プレビュー Canvas を各chに追加
        # ================================================================== #
        fg_outer = ttk.LabelFrame(control_frame, text="Function Generator (AO)",
                                  padding="3")
        fg_outer.pack(fill=tk.X, pady=(0, 3))

        # AO仕様ラベル
        ttk.Label(fg_outer,
                  text="Max 5 kS/s  |  ±10 V  |  Practical limit ~500 Hz",
                  font=('Arial', 7), foreground='gray').pack(anchor=tk.W)

        fg_cols = ttk.Frame(fg_outer)
        fg_cols.pack(fill=tk.X, pady=(4, 0))

        # AO0, AO1 それぞれのUIウィジェット変数を格納するリスト
        self.ao_waveform_vars  = []
        self.ao_freq_vars      = []
        self.ao_amp_vars       = []
        self.ao_offset_vars    = []
        self.ao_phase_vars     = []
        self.ao_preview_canvas = []   # 波形プレビュー用 tk.Canvas

        # プレビューCanvas サイズ（インスタンス変数として保持→描画時に直接参照）
        self._pv_w = 150
        self._pv_h = 120
        PV_W, PV_H = self._pv_w, self._pv_h

        AO_WAVEFORMS = ['Sine', 'Square', 'Triangle', 'Sawtooth', 'DC']
        AO_COLORS    = ['#E41A1C', '#377EB8']   # AO0:赤, AO1:青

        for ch_idx in range(2):
            col_frame = ttk.LabelFrame(fg_cols, text=f" AO{ch_idx} ",
                                       padding="6")
            col_frame.grid(row=0, column=ch_idx, padx=(0 if ch_idx == 0 else 4, 0),
                           sticky=tk.NSEW)
            fg_cols.columnconfigure(ch_idx, weight=1)

            # ── 左列：設定 ──────────────────────────── #
            left = ttk.Frame(col_frame)
            left.grid(row=0, column=0, sticky=tk.N)

            # 波形選択
            ttk.Label(left, text="Waveform",
                      font=('Arial', 8, 'bold')).grid(
                row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 1))
            wf_var = tk.StringVar(value='Sine')
            self.ao_waveform_vars.append(wf_var)
            wf_cb = ttk.Combobox(left, textvariable=wf_var,
                                 values=AO_WAVEFORMS, width=9, state='readonly')
            wf_cb.grid(row=1, column=0, columnspan=2, sticky=tk.EW, pady=(0, 3))

            # Frequency
            ttk.Label(left, text="Freq (Hz)",
                      font=('Arial', 8)).grid(row=2, column=0, sticky=tk.W)
            freq_var = tk.DoubleVar(value=10.0)
            self.ao_freq_vars.append(freq_var)
            ttk.Spinbox(left, textvariable=freq_var,
                        from_=0.1, to=500.0, increment=1.0,
                        width=7, format="%.1f").grid(
                row=2, column=1, sticky=tk.EW, padx=(2, 0))

            # Amplitude
            ttk.Label(left, text="Amp (V)",
                      font=('Arial', 8)).grid(row=3, column=0, sticky=tk.W, pady=(2, 0))
            amp_var = tk.DoubleVar(value=1.0)
            self.ao_amp_vars.append(amp_var)
            ttk.Spinbox(left, textvariable=amp_var,
                        from_=0.0, to=10.0, increment=0.1,
                        width=7, format="%.2f").grid(
                row=3, column=1, sticky=tk.EW, padx=(2, 0), pady=(2, 0))

            # Offset
            ttk.Label(left, text="Offset (V)",
                      font=('Arial', 8)).grid(row=4, column=0, sticky=tk.W, pady=(2, 0))
            off_var = tk.DoubleVar(value=0.0)
            self.ao_offset_vars.append(off_var)
            ttk.Spinbox(left, textvariable=off_var,
                        from_=-10.0, to=10.0, increment=0.1,
                        width=7, format="%.2f").grid(
                row=4, column=1, sticky=tk.EW, padx=(2, 0), pady=(2, 0))

            # Phase
            ttk.Label(left, text="Phase (°)",
                      font=('Arial', 8)).grid(row=5, column=0, sticky=tk.W, pady=(2, 0))
            phase_var = tk.DoubleVar(value=0.0)
            self.ao_phase_vars.append(phase_var)
            ttk.Spinbox(left, textvariable=phase_var,
                        from_=-360.0, to=360.0, increment=1.0,
                        width=7, format="%.1f").grid(
                row=5, column=1, sticky=tk.EW, padx=(2, 0), pady=(2, 0))

            # 出力ボタン
            out_btn = tk.Button(left,
                                text=f"▶ Output AO{ch_idx}",
                                font=('Arial', 9, 'bold'),
                                bg='#d9d9d9', fg='black',
                                relief=tk.RAISED,
                                command=lambda idx=ch_idx: self.toggle_ao_output(idx))
            out_btn.grid(row=6, column=0, columnspan=2,
                         sticky=tk.EW, pady=(6, 0))
            self.ao_out_btns[ch_idx] = out_btn

            # ── 右列：波形プレビュー Canvas ─────────── #
            pv = tk.Canvas(col_frame, width=PV_W, height=PV_H,
                           bg='white', highlightthickness=1,
                           highlightbackground='#aaaaaa')
            pv.grid(row=0, column=1, padx=(6, 0), sticky=tk.N)
            self.ao_preview_canvas.append(pv)

            # 設定変更時にプレビューを自動更新（デバウンス：前の予約をキャンセルしてから再予約）
            def _schedule_preview(idx=ch_idx, *_):
                if self.ao_preview_after[idx] is not None:
                    self.root.after_cancel(self.ao_preview_after[idx])
                self.ao_preview_after[idx] = self.root.after(
                    80, lambda i=idx: self._redraw_fg_preview(i))

            for var in (wf_var, freq_var, amp_var, off_var, phase_var):
                var.trace_add('write', _schedule_preview)
            wf_cb.bind('<<ComboboxSelected>>',
                       lambda e, idx=ch_idx: _schedule_preview(idx))

        # ================================================================== #
        # [New v2.1] Frequency Response (Bode) Measurement パネル
        # ================================================================== #
        fr_outer = ttk.LabelFrame(control_frame,
                                  text="Frequency Response (Bode)",
                                  padding="3")
        fr_outer.pack(fill=tk.X, pady=(0, 3))

        fr_grid = ttk.Frame(fr_outer)
        fr_grid.pack(fill=tk.X)
        for c in range(4):
            fr_grid.columnconfigure(c, weight=1)

        # 信号源候補（入力参照・応答に共通で使える）
        # AO0指令 + AI0〜AI7（ラベルは短縮）
        self.fr_source_options = ['AO0 cmd'] + \
            [f"AI{i}:{self.config['channels'][i]['name'][:6]}" for i in range(8)]

        # 入力参照（モータドライバ指令など）。デフォルトは AO0 指令値
        ttk.Label(fr_grid, text="Input ref:", font=('Arial', 8)).grid(
            row=0, column=0, sticky=tk.W)
        self.fr_input_var = tk.StringVar(value=self.fr_source_options[0])
        ttk.Combobox(fr_grid, textvariable=self.fr_input_var,
                     values=self.fr_source_options, width=10,
                     state='readonly').grid(
            row=0, column=1, columnspan=3, sticky=tk.EW, padx=2, pady=1)

        # 応答（出力）。デフォルトは AI0
        ttk.Label(fr_grid, text="Response:", font=('Arial', 8)).grid(
            row=1, column=0, sticky=tk.W)
        self.fr_output_var = tk.StringVar(value=self.fr_source_options[1])
        ttk.Combobox(fr_grid, textvariable=self.fr_output_var,
                     values=self.fr_source_options, width=10,
                     state='readonly').grid(
            row=1, column=1, columnspan=3, sticky=tk.EW, padx=2, pady=1)

        # 掃引パラメータ
        ttk.Label(fr_grid, text="f start (Hz):", font=('Arial', 8)).grid(
            row=2, column=0, sticky=tk.W)
        self.fr_fstart_var = tk.DoubleVar(value=1.0)
        ttk.Spinbox(fr_grid, textvariable=self.fr_fstart_var,
                    from_=0.1, to=500.0, increment=1.0, width=7,
                    format="%.1f").grid(row=2, column=1, sticky=tk.EW, padx=2)
        ttk.Label(fr_grid, text="f stop (Hz):", font=('Arial', 8)).grid(
            row=2, column=2, sticky=tk.W)
        self.fr_fstop_var = tk.DoubleVar(value=200.0)
        ttk.Spinbox(fr_grid, textvariable=self.fr_fstop_var,
                    from_=0.1, to=500.0, increment=10.0, width=7,
                    format="%.1f").grid(row=2, column=3, sticky=tk.EW, padx=2)

        ttk.Label(fr_grid, text="Points:", font=('Arial', 8)).grid(
            row=3, column=0, sticky=tk.W)
        self.fr_points_var = tk.IntVar(value=15)
        ttk.Spinbox(fr_grid, textvariable=self.fr_points_var,
                    from_=3, to=60, increment=1, width=7).grid(
            row=3, column=1, sticky=tk.EW, padx=2)
        ttk.Label(fr_grid, text="Amp (V):", font=('Arial', 8)).grid(
            row=3, column=2, sticky=tk.W)
        self.fr_amp_var = tk.DoubleVar(value=1.0)
        ttk.Spinbox(fr_grid, textvariable=self.fr_amp_var,
                    from_=0.1, to=10.0, increment=0.1, width=7,
                    format="%.2f").grid(row=3, column=3, sticky=tk.EW, padx=2)

        ttk.Label(fr_grid, text="Settle (s):", font=('Arial', 8)).grid(
            row=4, column=0, sticky=tk.W)
        self.fr_settle_var = tk.DoubleVar(value=0.5)
        ttk.Spinbox(fr_grid, textvariable=self.fr_settle_var,
                    from_=0.1, to=5.0, increment=0.1, width=7,
                    format="%.1f").grid(row=4, column=1, sticky=tk.EW, padx=2)
        ttk.Label(fr_grid, text="Meas (cyc):", font=('Arial', 8)).grid(
            row=4, column=2, sticky=tk.W)
        self.fr_cycles_var = tk.IntVar(value=10)
        ttk.Spinbox(fr_grid, textvariable=self.fr_cycles_var,
                    from_=2, to=50, increment=1, width=7).grid(
            row=4, column=3, sticky=tk.EW, padx=2)

        ttk.Label(fr_grid, text="Sweep:", font=('Arial', 8)).grid(
            row=5, column=0, sticky=tk.W)
        self.fr_logsweep_var = tk.StringVar(value='Log')
        ttk.Combobox(fr_grid, textvariable=self.fr_logsweep_var,
                     values=['Log', 'Linear'], width=7,
                     state='readonly').grid(row=5, column=1, sticky=tk.EW, padx=2)

        # 実行・状態表示ボタン
        self.fr_run_btn = tk.Button(
            fr_outer, text="▶ Measure Frequency Response",
            font=('Arial', 9, 'bold'), bg='#3366cc', fg='white',
            relief=tk.RAISED, command=self.toggle_freq_response)
        self.fr_run_btn.pack(fill=tk.X, pady=(5, 2))

        self.fr_status_label = ttk.Label(
            fr_outer, text="Ready", font=('Arial', 8), foreground='gray')
        self.fr_status_label.pack(anchor=tk.W)

        self.graph_frame = ttk.Frame(main_frame)
        self.graph_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 初回プレビュー描画（ウィジェット生成後に遅延実行）
        self.root.after(100, lambda: self._redraw_fg_preview(0))
        self.root.after(100, lambda: self._redraw_fg_preview(1))

        # [v2.4] DIO ボタンの初期表示とタスク構築（全lineデフォルトOUT）
        for _ln in range(8):
            self._update_dio_mode_button(_ln)
            self._update_dio_state_button(_ln)
        self.root.after(200, self._rebuild_dio_tasks)

    # ---------------------------------------------------------------------- #
    # グラフ初期設定
    # ---------------------------------------------------------------------- #
    def setup_graph(self):
        graph_cfg = self.config['graph']
        font_family = graph_cfg.get('font_family', 'Arial')

        plt.rcParams['font.family'] = font_family
        plt.rcParams['font.size'] = graph_cfg.get('font_size_tick', 10)

        self.fig, self.ax = plt.subplots(figsize=(12, 7))
        self.fig.set_facecolor('white')
        self.ax.set_facecolor('white')

        y_min = min(ch['y_min'] for ch in self.config['channels'])
        y_max = max(ch['y_max'] for ch in self.config['channels'])
        self.ax.set_ylim(y_min, y_max)
        self.ax.set_xlim(0, graph_cfg.get('time_window', 5.0))

        self.ax.set_xlabel(graph_cfg['xlabel'],
                           fontsize=graph_cfg['font_size_label'],
                           fontweight='bold', fontfamily=font_family)
        self.ax.set_ylabel(graph_cfg.get('ylabel', 'Voltage (V)'),
                           fontsize=graph_cfg['font_size_label'],
                           fontweight='bold', fontfamily=font_family)
        # -------------------------------------------------------------- #
        # グリッド描画：matplotlib の grid() / MinorLocator は
        # update_plot の draw_idle() と干渉するため使用しない。
        # 代わりに起動時に固定の Line2D オブジェクトとして描画し、
        # zorder=0 でデータラインの背面に固定する。
        # -------------------------------------------------------------- #
        if graph_cfg['grid']:
            x_min_g = 0.0
            x_max_g = graph_cfg.get('time_window', 5.0)
            y_min_g = min(ch['y_min'] for ch in self.config['channels'])
            y_max_g = max(ch['y_max'] for ch in self.config['channels'])
            x_range = x_max_g - x_min_g
            y_range = y_max_g - y_min_g

            # 主グリッド間隔（全体を5分割）
            x_major_step = x_range / 5.0
            y_major_step = y_range / 5.0
            # 補助グリッド間隔（主グリッドをさらに5分割）
            x_minor_step = x_major_step / 5.0
            y_minor_step = y_major_step / 5.0

            # 補助グリッド線（先に描画して背面へ）
            x_pos = x_min_g
            while x_pos <= x_max_g + 1e-9:
                self.ax.axvline(x_pos, color='gray', linewidth=0.4,
                                alpha=0.15, linestyle='-', zorder=0)
                x_pos += x_minor_step
            y_pos = y_min_g
            while y_pos <= y_max_g + 1e-9:
                self.ax.axhline(y_pos, color='gray', linewidth=0.4,
                                alpha=0.15, linestyle='-', zorder=0)
                y_pos += y_minor_step

            # 主グリッド線（補助の上、データラインの下）
            x_pos = x_min_g
            while x_pos <= x_max_g + 1e-9:
                self.ax.axvline(x_pos, color='gray', linewidth=0.8,
                                alpha=0.4, linestyle='--', zorder=1)
                x_pos += x_major_step
            y_pos = y_min_g
            while y_pos <= y_max_g + 1e-9:
                self.ax.axhline(y_pos, color='gray', linewidth=0.8,
                                alpha=0.4, linestyle='--', zorder=1)
                y_pos += y_major_step

        # 四方枠の描画（主軸）
        for spine in self.ax.spines.values():
            spine.set_visible(True)
            spine.set_linewidth(1.5)
            spine.set_color('black')

        # -------------------------------------------------------------- #
        # 第二Y軸ラベル（数値は主軸と完全共有、ラベルのみ独立）
        #
        # 手順:
        #   1. twinx() で ax2 を作成（データは描画しない）
        #   2. twinx() は主軸の ylabel を右側へ移動させようとするため、
        #      直後に ax の yaxis.set_label_position('left') で左に固定する
        #   3. ax2 の ylim・yticks を ax に合わせて同期
        #   4. ylim_changed コールバックで実行中も追従
        # -------------------------------------------------------------- #
        ylabel_right = graph_cfg.get('ylabel_right', '').strip()
        if ylabel_right:
            self.ax2 = self.ax.twinx()

            # twinx() 後に主軸ラベルが右へ動くのを明示的に防ぐ
            self.ax.yaxis.set_label_position('left')
            self.ax.yaxis.set_ticks_position('left')

            # ax2 の右軸設定
            self.ax2.set_ylim(self.ax.get_ylim())
            self.ax2.set_yticks(self.ax.get_yticks())
            self.ax2.set_ylim(self.ax.get_ylim())   # yticks 後に再同期
            self.ax2.set_ylabel(ylabel_right,
                                fontsize=graph_cfg['font_size_label'],
                                fontweight='bold', fontfamily=font_family)
            self.ax2.tick_params(axis='y',
                                 labelsize=graph_cfg['font_size_tick'],
                                 labelcolor='black')
            # ax2 の左 spine を非表示（主軸と重ならないように）
            self.ax2.spines['left'].set_visible(False)
            self.ax2.spines['right'].set_linewidth(1.5)
            self.ax2.spines['right'].set_color('black')
            # ax2 の上下 spine も主軸と揃える
            self.ax2.spines['top'].set_linewidth(1.5)
            self.ax2.spines['top'].set_color('black')
            self.ax2.spines['bottom'].set_visible(False)
        else:
            self.ax2 = None
            self.ax.yaxis.set_label_position('left')
            self.ax.yaxis.set_ticks_position('left')
            self.ax.tick_params(right=False, labelright=False)

        self.lines = []
        for i, ch in enumerate(self.config['channels']):
            line, = self.ax.plot([], [],
                                 linewidth=graph_cfg['line_width'],
                                 color=COLORS[i],
                                 label=ch['name'],
                                 alpha=0.8,
                                 animated=False)
            self.lines.append(line)

        self.update_legend()
        self.ax.set_title(graph_cfg['title'],
                          fontsize=graph_cfg['font_size_title'],
                          fontweight='bold', fontfamily=font_family, pad=15)
        # 右Y軸ラベルがある場合は右側に余白を確保
        ylabel_right = graph_cfg.get('ylabel_right', '').strip()
        right_margin = 0.93 if ylabel_right else 1.0
        self.fig.tight_layout(pad=2.0, rect=[0.02, 0, right_margin, 1])

        # ax2 の ylim を主軸と常に同期させるコールバック
        # _syncing フラグで ax2→ax→ax2 の無限ループを防止する
        if hasattr(self, 'ax2') and self.ax2 is not None:
            self._ax2_syncing = False
            def _sync_ax2(ax):
                if self._ax2_syncing:
                    return
                self._ax2_syncing = True
                try:
                    lo, hi = ax.get_ylim()
                    self.ax2.set_ylim(lo, hi)
                    self.ax2.set_yticks(ax.get_yticks())
                    self.ax2.set_ylim(lo, hi)
                finally:
                    self._ax2_syncing = False
            self.ax.callbacks.connect('ylim_changed', _sync_ax2)

        self.canvas = FigureCanvasTkAgg(self.fig, master=self.graph_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        toolbar_frame = ttk.Frame(self.graph_frame)
        toolbar_frame.pack(fill=tk.X)
        NavigationToolbar2Tk(self.canvas, toolbar_frame).update()

        self.setup_context_menu()

    def update_legend(self):
        if hasattr(self, 'legend') and self.legend:
            self.legend.remove()

        active_lines, active_labels = [], []
        for i, ch in enumerate(self.config['channels']):
            if self.active_channels[i]:
                active_lines.append(self.lines[i])
                active_labels.append(ch['name'])

        if active_lines:
            graph_cfg = self.config['graph']
            self.legend = self.ax.legend(
                active_lines, active_labels,
                loc=graph_cfg.get('legend_position', 'upper right'),
                fontsize=graph_cfg['font_size_legend'],
                prop={'family': graph_cfg.get('font_family', 'Arial')},
                framealpha=0.9, edgecolor='black',
                ncol=min(graph_cfg.get('legend_columns', 2), len(active_lines)),
                columnspacing=1.0)

    # ---------------------------------------------------------------------- #
    # 右クリックメニュー
    # ---------------------------------------------------------------------- #
    def setup_context_menu(self):
        self.context_menu = tk.Menu(self.root, tearoff=0)

        copy_menu = tk.Menu(self.context_menu, tearoff=0)
        copy_menu.add_command(label="PNG (High Quality)",
                              command=lambda: self.copy_to_clipboard('png'))
        copy_menu.add_command(label="JPEG (Compressed)",
                              command=lambda: self.copy_to_clipboard('jpeg'))
        copy_menu.add_separator()
        copy_menu.add_command(label="PNG (Screen Quality)",
                              command=lambda: self.copy_to_clipboard('png', dpi=100))
        self.context_menu.add_cascade(label="📋 Copy as...", menu=copy_menu)

        save_menu = tk.Menu(self.context_menu, tearoff=0)
        for fmt in ['png', 'jpeg', 'tiff', 'pdf', 'svg']:
            label_map = {'png': 'PNG (High-Res)', 'jpeg': 'JPEG',
                         'tiff': 'TIFF', 'pdf': 'PDF (Vector)', 'svg': 'SVG (Vector)'}
            save_menu.add_command(label=label_map[fmt],
                                  command=lambda f=fmt: self.save_figure_format(f))
        self.context_menu.add_cascade(label="💾 Save as...", menu=save_menu)

        self.canvas.get_tk_widget().bind("<Button-3>", self.show_context_menu)

    def show_context_menu(self, event):
        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()

    # ---------------------------------------------------------------------- #
    # クリップボード・保存
    # ---------------------------------------------------------------------- #
    def copy_to_clipboard(self, format='png', dpi=150):
        try:
            self.canvas.draw()
            self.root.update()
            buf = io.BytesIO()
            self.fig.savefig(buf, format=format, dpi=dpi, bbox_inches='tight',
                             facecolor='white', edgecolor='none')
            buf.seek(0)
            img = Image.open(buf)
            try:
                import win32clipboard
                output = io.BytesIO()
                img.convert('RGB').save(output, 'BMP')
                data = output.getvalue()[14:]
                output.close()
                win32clipboard.OpenClipboard()
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
                win32clipboard.CloseClipboard()
                messagebox.showinfo("Success",
                                    f"Graph copied to clipboard as {format.upper()}!")
            except ImportError:
                messagebox.showinfo(
                    "Info",
                    "pywin32 module recommended for clipboard functionality.\n"
                    "Install with: pip install pywin32")
            buf.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to copy to clipboard:\n{e}")

    def save_figure_format(self, format):
        format_extensions = {
            'png': [("PNG Image", "*.png")],
            'jpeg': [("JPEG Image", "*.jpg"), ("JPEG Image", "*.jpeg")],
            'tiff': [("TIFF Image", "*.tiff"), ("TIFF Image", "*.tif")],
            'pdf': [("PDF Document", "*.pdf")],
            'svg': [("SVG Vector", "*.svg")]
        }
        default_ext = {'png': '.png', 'jpeg': '.jpg', 'tiff': '.tiff',
                       'pdf': '.pdf', 'svg': '.svg'}
        filename = filedialog.asksaveasfilename(
            defaultextension=default_ext[format],
            filetypes=format_extensions[format],
            initialfile=f"Figure_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}")
        if filename:
            dpi = 300 if format in ['png', 'jpeg', 'tiff'] else None
            self.fig.savefig(filename, dpi=dpi, bbox_inches='tight',
                             facecolor='white', edgecolor='none')
            messagebox.showinfo("Success", f"Figure saved!\n{filename}")

    def quick_export_menu(self):
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Copy as PNG",
                         command=lambda: self.copy_to_clipboard('png'))
        menu.add_command(label="Copy as JPEG",
                         command=lambda: self.copy_to_clipboard('jpeg'))
        menu.add_separator()
        menu.add_command(label="Save as PNG...",
                         command=lambda: self.save_figure_format('png'))
        menu.add_command(label="Save as PDF...",
                         command=lambda: self.save_figure_format('pdf'))
        btn = self.root.focus_get()
        x = btn.winfo_rootx()
        y = btn.winfo_rooty() + btn.winfo_height()
        menu.tk_popup(x, y)

    def clear_all_data(self):
        if not messagebox.askyesno("Confirm",
                                   "Clear all channel data?\n"
                                   "This will reset all graph data.",
                                   icon='warning'):
            return
        try:
            for i in range(8):
                self.display_buffer[i].clear()
                self.lines[i].set_data([], [])
                self.current_values[i] = 0.0
                self.value_labels[i].config(
                    text=f"+0.000 {self.config['channels'][i]['unit']}")
            self.canvas.draw_idle()
            messagebox.showinfo("Success", "All channel data has been cleared.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to clear data:\n{e}")

    # ---------------------------------------------------------------------- #
    # チャンネル操作
    # ---------------------------------------------------------------------- #
    def update_channel_name(self, channel_idx):
        new_name = self.channel_name_vars[channel_idx].get()
        self.config['channels'][channel_idx]['name'] = new_name
        self.lines[channel_idx].set_label(new_name)
        self.update_legend()
        self.canvas.draw_idle()

    def update_channel_factor(self, channel_idx):
        try:
            new_factor = self.channel_factor_vars[channel_idx].get()
            self.config['channels'][channel_idx]['conversion_factor'] = new_factor
        except tk.TclError:
            self.channel_factor_vars[channel_idx].set(
                self.config['channels'][channel_idx]['conversion_factor'])

    def toggle_offset(self, channel_idx):
        """Zero/Reset トグル：未設定なら現在値をゼロ点に、設定済みならリセット"""
        btn = self.zero_reset_btns[channel_idx]
        unit = self.config['channels'][channel_idx]['unit']

        if self.offsets[channel_idx] == 0.0:
            # ── Zero 操作 ──────────────────────────────────
            raw = self.current_values[channel_idx]
            self.offsets[channel_idx] = raw
            self.offset_labels[channel_idx].config(
                text=f"{raw:+.4f} {unit}", foreground='#E41A1C')
            btn.config(text="Reset")
        else:
            # ── Reset 操作 ─────────────────────────────────
            self.offsets[channel_idx] = 0.0
            self.offset_labels[channel_idx].config(
                text="---", foreground='gray')
            btn.config(text="Zero")

        # どちらの操作でもバッファをクリアして新しい基準でグラフ再スタート
        self.display_buffer[channel_idx].clear()

    # 後方互換用エイリアス（他所から呼ばれている場合のため残す）
    def set_offset(self, channel_idx):
        if self.offsets[channel_idx] == 0.0:
            self.toggle_offset(channel_idx)

    def reset_offset(self, channel_idx):
        if self.offsets[channel_idx] != 0.0:
            self.toggle_offset(channel_idx)

    def update_active_channels(self):
        self.active_channels = [var.get() for var in self.channel_vars]
        # キャッシュを更新
        self._active_ch_indices = [
            i for i, a in enumerate(self.active_channels) if a]
        for i, active in enumerate(self.active_channels):
            self.lines[i].set_alpha(0.8 if active else 0.1)
            if not active:
                self.lines[i].set_data([], [])
        self.update_legend()
        self.canvas.draw_idle()

    def open_settings(self):
        editor = ConfigEditor(self.root, self.config)
        self.root.wait_window(editor)
        if editor.result:
            self.config = editor.result
            self.canvas.get_tk_widget().destroy()
            self.setup_graph()

    # ---------------------------------------------------------------------- #
    # 録画
    # ---------------------------------------------------------------------- #
    def start_recording(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Recording Settings")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Recording Duration (seconds):",
                  font=('Arial', 11)).pack(pady=10)
        duration_var = tk.DoubleVar(value=10.0)
        duration_entry = ttk.Entry(dialog, textvariable=duration_var,
                                   font=('Arial', 12), width=15)
        duration_entry.pack(pady=5)
        duration_entry.focus()

        def start():
            try:
                duration = duration_var.get()
                if duration <= 0:
                    raise ValueError("Duration must be positive")
                self.target_samples = int(
                    duration * self.config['device']['sampling_rate'])
                self.recording_count = 0
                num_active = sum(self.active_channels)
                self.recorded_data = [[] for _ in range(num_active)]
                self.is_recording = True
                self.record_btn.config(state=tk.DISABLED)
                self.status_label.config(text="Recording...", foreground='red')
                dialog.destroy()
            except ValueError as e:
                messagebox.showerror("Error", f"Invalid input:\n{e}")

        ttk.Button(dialog, text="Start", command=start).pack(pady=10)
        ttk.Button(dialog, text="Cancel", command=dialog.destroy).pack()

    # ---------------------------------------------------------------------- #
    # DAQ タスク
    # ---------------------------------------------------------------------- #
    def start_task(self):
        try:
            self.task = nidaqmx.Task()
            for i, (ch, active) in enumerate(
                    zip(self.config['channels'], self.active_channels)):
                if active:
                    channel_name = (f"{self.config['device']['device_name']}"
                                    f"/ai{ch['id']}")
                    term_config = TerminalConfiguration.RSE
                    if self.config['device']['terminal_config'] == "DIFF":
                        term_config = TerminalConfiguration.DIFFERENTIAL
                    elif self.config['device']['terminal_config'] == "NRSE":
                        term_config = TerminalConfiguration.NRSE
                    self.task.ai_channels.add_ai_voltage_chan(
                        channel_name,
                        terminal_config=term_config,
                        min_val=-10.0, max_val=10.0)

            self.task.timing.cfg_samp_clk_timing(
                rate=self.config['device']['sampling_rate'],
                sample_mode=AcquisitionType.CONTINUOUS,
                samps_per_chan=self.config['device']['samples_per_channel'])
            self.task.start()

            buffer_size = int(
                self.config['graph'].get('time_window', 5.0) *
                self.config['device']['sampling_rate'])
            self.display_buffer = [
                collections.deque(maxlen=buffer_size) for _ in range(8)]

        except Exception as e:
            messagebox.showerror("Error", f"Failed to start DAQ task:\n{e}")

    # ---------------------------------------------------------------------- #
    # [Fix 3] Hold/Resume: Hold 中もバッファを読み捨てる
    # ---------------------------------------------------------------------- #
    def update_plot(self):
        """グラフの更新（20 ms 周期）"""

        # [v2.1] 周波数特性測定中はメインループをスキップ（AI競合回避）
        if self.fr_running:
            self.update_id = self.root.after(50, self.update_plot)
            return

        if self.is_paused:
            try:
                self.task.read(number_of_samples_per_channel=READ_ALL_AVAILABLE)
            except Exception:
                pass
            self.update_id = self.root.after(20, self.update_plot)
            return

        try:
            data = self.task.read(
                number_of_samples_per_channel=READ_ALL_AVAILABLE)

            if not data or (isinstance(data, list) and len(data[0]) == 0):
                self.update_id = self.root.after(20, self.update_plot)
                return

            if not isinstance(data[0], list):
                data = [[d] for d in data]

            num_new = len(data[0])
            active_ch_indices = self._active_ch_indices

            for data_idx, ch_idx in enumerate(active_ch_indices):
                cf     = self.config['channels'][ch_idx]['conversion_factor']
                offset = self.offsets[ch_idx]
                raw = np.asarray(data[data_idx], dtype=np.float64)
                converted = raw * cf
                self.display_buffer[ch_idx].extend((converted - offset).tolist())
                self.current_values[ch_idx] = float(converted[-1])

            if self.is_recording:
                for data_idx in range(len(active_ch_indices)):
                    self.recorded_data[data_idx].extend(data[data_idx])
                self.recording_count += num_new
                progress = min(
                    100,
                    int(self.recording_count / self.target_samples * 100))
                self.progress_label.config(
                    text=(f"Progress: {progress}% "
                          f"({self.recording_count}/{self.target_samples} samples)"))
                if self.recording_count >= self.target_samples:
                    self.finish_recording()

            self.update_counter += 1
            if self.update_counter >= self.draw_interval:
                self.update_counter = 0
                time_window = self.config['graph'].get('time_window', 5.0)
                for ch_idx in active_ch_indices:
                    buf = self.display_buffer[ch_idx]
                    n = len(buf)
                    if n:
                        # time軸はバッファ長が変わった時だけ再生成（キャッシュ）
                        if self._time_axis_cache_len != n:
                            self._time_axis_cache = np.linspace(0, time_window, n)
                            self._time_axis_cache_len = n
                        self.lines[ch_idx].set_data(
                            self._time_axis_cache, np.fromiter(buf, dtype=np.float64, count=n))
                        # 値ラベルは描画タイミングのみ更新（毎フレーム不要）
                        ch_config = self.config['channels'][ch_idx]
                        self.value_labels[ch_idx].config(
                            text=f"{self.current_values[ch_idx] - self.offsets[ch_idx]:+.3f} {ch_config['unit']}")
                self.canvas.draw_idle()
                # [v2.4] DI監視：描画タイミングのついでに軽量読み取り
                self._read_dio_in()

        except Exception as e:
            print(f"Error in update_plot: {e}")

        self.update_id = self.root.after(20, self.update_plot)

    # ---------------------------------------------------------------------- #
    def finish_recording(self):
        self.is_recording = False
        self.record_btn.config(state=tk.NORMAL)
        if not self.is_paused:
            self.status_label.config(text="Monitoring...", foreground='green')
        self.progress_label.config(text="")
        self.save_data()

    # ---------------------------------------------------------------------- #
    # [Fix 3] Hold / Resume
    # ---------------------------------------------------------------------- #
    def toggle_hold(self):
        """Hold と Resume を切り替える"""
        self.is_paused = not self.is_paused

        if self.is_paused:
            self.hold_btn.config(text="▶ Resume")
            self.status_label.config(text="Hold", foreground='orange')
            # -------------------------------------------------------------- #
            # [New v1.5] Hold Snapshot: Hold 押下時点の表示バッファをExcel保存
            # -------------------------------------------------------------- #
            self._save_hold_snapshot()
        else:
            # Resume: 表示バッファをクリアして再スタート（古いデータが混入しない）
            for buf in self.display_buffer:
                buf.clear()
            self.hold_btn.config(text="⏸ Hold")
            self.status_label.config(text="Monitoring...", foreground='green')

    # ---------------------------------------------------------------------- #
    # [New v1.5 / Updated v1.8] Hold Snapshot 保存
    # ---------------------------------------------------------------------- #
    def _save_hold_snapshot(self):
        """Hold 押下時点の display_buffer を Excel へ保存する。

        Sheet1 "Data"  : Time[s] + 各チャンネルの変換・オフセット済み値
        Sheet2 "Graph" : 画面表示と同じスタイルのグラフ画像を貼り付け
        """
        try:
            active_ch_indices = [
                i for i, active in enumerate(self.active_channels) if active]
            if not active_ch_indices:
                return

            # ── データ収集 ─────────────────────────────────────── #
            df_data = {}
            n_samples = 0
            for ch_idx in active_ch_indices:
                ch_config = self.config['channels'][ch_idx]
                buf = list(self.display_buffer[ch_idx])
                col_name = f"{ch_config['name']}_{ch_config['unit']}"
                df_data[col_name] = buf
                n_samples = max(n_samples, len(buf))

            if n_samples == 0:
                return

            for key in df_data:
                if len(df_data[key]) < n_samples:
                    df_data[key] += [float('nan')] * (n_samples - len(df_data[key]))

            sampling_rate = self.config['device']['sampling_rate']
            t_axis = [t / sampling_rate for t in range(n_samples)]

            df = pd.DataFrame(df_data)
            df.insert(0, "Time[s]", t_axis)
            df = df.round(6)

            # ── グラフ画像生成（画面と同スタイル）─────────────── #
            graph_cfg  = self.config['graph']
            font_family = graph_cfg.get('font_family', 'Arial')

            fig_snap, ax_snap = plt.subplots(figsize=(12, 7))
            fig_snap.set_facecolor('white')
            ax_snap.set_facecolor('white')

            # 軸範囲・ラベルを live plot と同じに設定
            y_min = min(ch['y_min'] for ch in self.config['channels'])
            y_max = max(ch['y_max'] for ch in self.config['channels'])
            ax_snap.set_xlim(0, graph_cfg.get('time_window', 5.0))
            ax_snap.set_ylim(y_min, y_max)
            ax_snap.set_xlabel(graph_cfg['xlabel'],
                               fontsize=graph_cfg['font_size_label'],
                               fontweight='bold', fontfamily=font_family)
            ax_snap.set_ylabel(graph_cfg.get('ylabel', 'Voltage (V)'),
                               fontsize=graph_cfg['font_size_label'],
                               fontweight='bold', fontfamily=font_family)
            ax_snap.set_title(graph_cfg['title'] + "  [Hold Snapshot]",
                              fontsize=graph_cfg['font_size_title'],
                              fontweight='bold', fontfamily=font_family, pad=15)

            # グリッド（live plot と同じ固定Line2D方式）
            if graph_cfg['grid']:
                x_range = graph_cfg.get('time_window', 5.0)
                y_range = y_max - y_min
                x_maj = x_range / 5.0;  x_min_s = x_maj / 5.0
                y_maj = y_range / 5.0;  y_min_s = y_maj / 5.0
                xp = 0.0
                while xp <= x_range + 1e-9:
                    ax_snap.axvline(xp, color='gray', linewidth=0.4,
                                   alpha=0.15, zorder=0)
                    xp += x_min_s
                yp = y_min
                while yp <= y_max + 1e-9:
                    ax_snap.axhline(yp, color='gray', linewidth=0.4,
                                   alpha=0.15, zorder=0)
                    yp += y_min_s
                xp = 0.0
                while xp <= x_range + 1e-9:
                    ax_snap.axvline(xp, color='gray', linewidth=0.8,
                                   alpha=0.4, linestyle='--', zorder=1)
                    xp += x_maj
                yp = y_min
                while yp <= y_max + 1e-9:
                    ax_snap.axhline(yp, color='gray', linewidth=0.8,
                                   alpha=0.4, linestyle='--', zorder=1)
                    yp += y_maj

            # 枠線
            for spine in ax_snap.spines.values():
                spine.set_visible(True)
                spine.set_linewidth(1.5)
                spine.set_color('black')

            ax_snap.tick_params(labelsize=graph_cfg['font_size_tick'])

            # 第二Y軸ラベル
            ylabel_right = graph_cfg.get('ylabel_right', '').strip()
            if ylabel_right:
                ax2_snap = ax_snap.twinx()
                ax_snap.yaxis.set_label_position('left')
                ax_snap.yaxis.set_ticks_position('left')
                ax2_snap.set_ylim(ax_snap.get_ylim())
                ax2_snap.set_yticks(ax_snap.get_yticks())
                ax2_snap.set_ylim(ax_snap.get_ylim())
                ax2_snap.set_ylabel(ylabel_right,
                                    fontsize=graph_cfg['font_size_label'],
                                    fontweight='bold', fontfamily=font_family)
                ax2_snap.tick_params(axis='y',
                                     labelsize=graph_cfg['font_size_tick'])
                ax2_snap.spines['left'].set_visible(False)
                ax2_snap.spines['right'].set_linewidth(1.5)

            # データプロット（live plot と同じ色・線幅）
            plot_lines = []
            plot_labels = []
            for ch_idx in active_ch_indices:
                ch_config = self.config['channels'][ch_idx]
                col_name  = f"{ch_config['name']}_{ch_config['unit']}"
                y_data    = df[col_name].values
                x_data    = df["Time[s]"].values
                line, = ax_snap.plot(x_data, y_data,
                                     linewidth=graph_cfg['line_width'],
                                     color=COLORS[ch_idx],
                                     alpha=0.8,
                                     label=ch_config['name'])
                plot_lines.append(line)
                plot_labels.append(ch_config['name'])

            # 凡例
            if plot_lines:
                ax_snap.legend(
                    plot_lines, plot_labels,
                    loc=graph_cfg.get('legend_position', 'upper right'),
                    fontsize=graph_cfg['font_size_legend'],
                    prop={'family': font_family},
                    framealpha=0.9, edgecolor='black',
                    ncol=min(graph_cfg.get('legend_columns', 2),
                             len(plot_lines)),
                    columnspacing=1.0)

            right_margin = 0.93 if ylabel_right else 1.0
            fig_snap.tight_layout(pad=2.0, rect=[0.02, 0, right_margin, 1])

            # PNG バイト列として取得
            img_buf = io.BytesIO()
            fig_snap.savefig(img_buf, format='png', dpi=150,
                             bbox_inches='tight',
                             facecolor='white', edgecolor='none')
            img_buf.seek(0)
            plt.close(fig_snap)

            # ── Excel 書き出し（openpyxl）──────────────────────── #
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            # --- Sheet1: Data ---
            ws_data = wb.active
            ws_data.title = "Data"

            # ヘッダー行スタイル
            hdr_fill = PatternFill("solid", fgColor="1F4E79")
            hdr_font = Font(name=font_family, bold=True, color="FFFFFF", size=11)
            hdr_align = Alignment(horizontal="center", vertical="center")
            thin = Side(style='thin', color='AAAAAA')
            hdr_border = Border(left=thin, right=thin, bottom=thin)

            cols = ["Time[s]"] + list(df_data.keys())
            for c_idx, col_name in enumerate(cols, start=1):
                cell = ws_data.cell(row=1, column=c_idx, value=col_name)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = hdr_align
                cell.border = hdr_border
                ws_data.column_dimensions[
                    get_column_letter(c_idx)].width = max(14, len(col_name) + 2)

            ws_data.row_dimensions[1].height = 20

            # データ行（交互背景色）
            fill_odd  = PatternFill("solid", fgColor="EBF3FB")
            fill_even = PatternFill("solid", fgColor="FFFFFF")
            data_font = Font(name='Courier New', size=10)
            num_align = Alignment(horizontal="right")
            data_border = Border(left=thin, right=thin)

            all_rows = [t_axis] + [df_data[k] for k in df_data]
            for r_idx in range(n_samples):
                fill = fill_odd if r_idx % 2 == 0 else fill_even
                for c_idx in range(len(cols)):
                    val = all_rows[c_idx][r_idx]
                    cell = ws_data.cell(row=r_idx + 2, column=c_idx + 1,
                                        value=round(val, 6) if val == val else None)
                    cell.font = data_font
                    cell.fill = fill
                    cell.alignment = num_align
                    cell.border = data_border

            # ウィンドウ枠の固定（ヘッダー行）
            ws_data.freeze_panes = "A2"

            # --- Sheet2: Graph ---
            ws_graph = wb.create_sheet(title="Graph")
            xl_img = XLImage(img_buf)
            # Excel上のサイズ調整（dpi=150、figsize=(12,7) → 約 1800×1050 px）
            xl_img.width  = 900
            xl_img.height = 525
            ws_graph.add_image(xl_img, "B2")

            # タイムスタンプをグラフシートのA1に記録
            now_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            ws_graph["A1"] = f"Hold Snapshot  {now_str}"
            ws_graph["A1"].font = Font(bold=True, size=11)

            # ── 保存 ───────────────────────────────────────────── #
            save_dir = r"C:\PRG\data"
            os.makedirs(save_dir, exist_ok=True)
            save_path = os.path.join(save_dir, f"Hold_Snapshot_{now_str}.xlsx")
            wb.save(save_path)

            self.status_label.config(
                text="Hold  ✓ Snapshot saved", foreground='orange')

        except Exception as e:
            print(f"[Hold Snapshot] Save error: {e}")
            import traceback; traceback.print_exc()

    # ---------------------------------------------------------------------- #
    # データ保存
    # ---------------------------------------------------------------------- #
    def save_data(self):
        try:
            active_ch_indices = [
                i for i, active in enumerate(self.active_channels) if active]
            df_data = {}
            for data_idx, ch_idx in enumerate(active_ch_indices):
                ch_config = self.config['channels'][ch_idx]
                voltage_data = self.recorded_data[data_idx][:self.target_samples]
                df_data[f"{ch_config['name']}_V"] = voltage_data
                df_data[f"{ch_config['name']}_{ch_config['unit']}"] = [
                    v * ch_config['conversion_factor'] for v in voltage_data]
            df = pd.DataFrame(df_data)
            t_axis = [t / self.config['device']['sampling_rate']
                      for t in range(self.target_samples)]
            df.insert(0, "Time[s]", t_axis)
            df = df.round(6)

            save_dir = r"C:\PRG\data"
            if not os.path.exists(save_dir):
                os.makedirs(save_dir)
            now_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            save_path = os.path.join(save_dir, f"USB6002_Data_{now_str}.xlsx")
            df.to_excel(save_path, index=False)
            messagebox.showinfo("Success", f"Data saved successfully!\n{save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save data:\n{e}")

    def save_figure(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[
                ("PNG Image (High-Res)", "*.png"),
                ("JPEG Image", "*.jpg"),
                ("TIFF Image", "*.tiff"),
                ("PDF Vector", "*.pdf"),
                ("SVG Vector", "*.svg")],
            initialfile=f"Figure_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}")
        if filename:
            self.fig.savefig(filename, dpi=300, bbox_inches='tight',
                             facecolor='white', edgecolor='none')
            messagebox.showinfo("Success", f"Figure saved!\n{filename}")

    # ---------------------------------------------------------------------- #
    # 終了処理
    # ---------------------------------------------------------------------- #
    def cleanup(self):
        if self.update_id:
            self.root.after_cancel(self.update_id)
            self.update_id = None
        if self.task:
            try:
                self.task.stop()
                self.task.close()
            except Exception:
                pass
            self.task = None
        # [New v1.6] AO タスクも終了
        for i in range(2):
            self._stop_ao(i)
        # [New v2.4] Digital I/O タスクも終了（DO全LOW化してから解放）
        self._stop_dio()

    # ================================================================== #
    # [New v1.7] Function Generator 波形プレビュー描画
    # ================================================================== #

    def _redraw_fg_preview(self, ch_idx):
        """波形プレビュー Canvas を現在の設定値で再描画する。"""
        import math as _math
        self.ao_preview_after[ch_idx] = None
        pv = self.ao_preview_canvas[ch_idx]
        W  = self._pv_w
        H  = self._pv_h
        pv.delete('all')

        try:
            waveform  = self.ao_waveform_vars[ch_idx].get()
            freq      = max(0.01, self.ao_freq_vars[ch_idx].get())
            amp       = float(self.ao_amp_vars[ch_idx].get())
            offset    = float(self.ao_offset_vars[ch_idx].get())
            phase_deg = float(self.ao_phase_vars[ch_idx].get())
        except Exception:
            return

        AO_COLORS = ['#cc0000', '#0055cc']
        wave_color = AO_COLORS[ch_idx]
        grid_color = '#dddddd'
        axis_color = '#888888'
        text_color = '#333333'
        clip_color = '#dd2200'

        # ── レイアウト定数 ──────────────────────────────────────── #
        margin_l   = 34
        margin_r   = 4
        margin_top = 22
        margin_bot = 6
        plot_w = W - margin_l - margin_r
        plot_h = H - margin_top - margin_bot

        # ── 波形サンプル生成（2周期）────────────────────────────── #
        n_pts = 300
        n_periods = 2.5
        t = np.linspace(0, n_periods, n_pts, endpoint=False)
        phi = np.deg2rad(phase_deg)

        if waveform == 'DC':
            y_vals = np.full(n_pts, offset)
        elif waveform == 'Sine':
            y_vals = amp * np.sin(2 * np.pi * t + phi) + offset
        elif waveform == 'Square':
            sq = np.sign(np.sin(2 * np.pi * t + phi))
            sq[sq == 0] = 1.0
            y_vals = amp * sq + offset
        elif waveform == 'Triangle':
            y_vals = amp * (2 / np.pi) * np.arcsin(
                np.clip(np.sin(2 * np.pi * t + phi), -1.0, 1.0)) + offset
        elif waveform == 'Sawtooth':
            t_shift = (t + phase_deg / 360.0) % 1.0
            y_vals = amp * (2 * t_shift - 1) + offset
        else:
            y_vals = np.zeros(n_pts)

        y_max_raw = float(np.max(y_vals))
        y_min_raw = float(np.min(y_vals))
        clipped   = bool(np.any(y_vals > 10.0) or np.any(y_vals < -10.0))
        y_vals    = np.clip(y_vals, -10.0, 10.0)

        # ── y軸表示範囲（振幅に合わせて自動スケール、±10V上限）── #
        if waveform == 'DC':
            y_half = 1.0
        else:
            y_half = max(amp * 1.2, 0.2)
        y_lo = max(-10.0, offset - y_half)
        y_hi = min( 10.0, offset + y_half)
        if y_hi - y_lo < 0.01:
            y_lo -= 0.5; y_hi += 0.5
        y_span = y_hi - y_lo

        # 座標変換（numpy ベクトル演算用）
        def v2y(v):
            return margin_top + plot_h * (1.0 - (float(v) - y_lo) / y_span)

        # ── グリッド線 ───────────────────────────────────────────── #
        raw_step = y_span / 4.0
        if raw_step > 0:
            mag = 10 ** _math.floor(_math.log10(raw_step))
            grid_step = next(
                (mag * f for f in (1, 2, 2.5, 5, 10) if mag * f >= raw_step),
                raw_step)
        else:
            grid_step = 1.0

        yg = _math.ceil(y_lo / grid_step) * grid_step
        while yg <= y_hi + 1e-9:
            cy    = v2y(yg)
            color = axis_color if abs(yg) < grid_step * 0.01 else grid_color
            dash  = () if abs(yg) < grid_step * 0.01 else (2, 3)
            pv.create_line(margin_l, cy, W - margin_r, cy,
                           fill=color, width=1, dash=dash)
            pv.create_text(margin_l - 2, cy, text=f"{yg:.2g}",
                           anchor=tk.E, fill=text_color, font=('Arial', 7))
            yg = round(yg + grid_step, 10)

        # ── 波形ポリライン（numpy で一括変換）───────────────────── #
        xs = margin_l + (np.arange(n_pts) / n_pts) * plot_w
        ys = margin_top + plot_h * (1.0 - (y_vals - y_lo) / y_span)
        # Canvas の描画領域にクリップ（はみ出し防止）
        ys = np.clip(ys, margin_top, margin_top + plot_h)
        coords = np.empty(n_pts * 2)
        coords[0::2] = xs
        coords[1::2] = ys
        smooth = waveform in ('Sine', 'Triangle')
        pv.create_line(*coords.tolist(),
                       fill=clip_color if clipped else wave_color,
                       width=2, smooth=smooth, joinstyle=tk.ROUND)

        # ── オフセット破線 ───────────────────────────────────────── #
        if waveform != 'DC' and y_lo <= offset <= y_hi:
            yo = v2y(offset)
            pv.create_line(margin_l, yo, W - margin_r, yo,
                           fill='#cc8800', width=1, dash=(4, 3))

        # ── 注釈テキスト ─────────────────────────────────────────── #
        if waveform == 'DC':
            info_top = f"DC  {offset:+.2f}V"
        else:
            info_top = f"{waveform}  {freq:.1f}Hz"
        pv.create_text(margin_l + 2, 3, text=info_top,
                       anchor=tk.NW, fill=text_color, font=('Courier', 9, 'bold'))

        if waveform != 'DC':
            pv.create_text(W - margin_r - 2, 3,
                           text=f"A={amp:.2f}V φ={phase_deg:.0f}°",
                           anchor=tk.NE, fill=text_color, font=('Courier', 9, 'bold'))

        # ピーク電圧ラベル
        if y_lo <= y_max_raw <= y_hi:
            pv.create_text(margin_l + 2, v2y(y_max_raw) - 1,
                           text=f"{y_max_raw:+.2f}V", anchor=tk.SW,
                           fill=wave_color, font=('Courier', 8, 'bold'))
        if y_lo <= y_min_raw <= y_hi and abs(y_max_raw - y_min_raw) > y_span * 0.05:
            pv.create_text(margin_l + 2, v2y(y_min_raw) + 1,
                           text=f"{y_min_raw:+.2f}V", anchor=tk.NW,
                           fill=wave_color, font=('Courier', 8, 'bold'))

        if clipped:
            pv.create_text(W // 2, H - 2, text="⚠ CLIPPED (>±10V)",
                           anchor=tk.S, fill=clip_color, font=('Arial', 8, 'bold'))


    # ================================================================== #
    # [New v1.6] Function Generator ロジック
    # ================================================================== #

    # AO サンプリングレート（USB-6002 仕様上限）
    AO_RATE = 5000

    def _make_waveform(self, ch_idx):
        """指定 AO ch の設定から1周期分の波形配列を生成して返す。
        DC の場合はオフセット値の単一要素配列を返す。
        """
        waveform = self.ao_waveform_vars[ch_idx].get()
        freq     = max(0.01, self.ao_freq_vars[ch_idx].get())
        amp      = self.ao_amp_vars[ch_idx].get()
        offset   = self.ao_offset_vars[ch_idx].get()
        phase_deg = self.ao_phase_vars[ch_idx].get()

        if waveform == 'DC':
            return np.array([np.clip(offset, -10.0, 10.0)])

        # 1周期分のサンプル数（最低4点）
        n = max(4, int(round(self.AO_RATE / freq)))
        t = np.linspace(0, 1, n, endpoint=False)
        phi = np.deg2rad(phase_deg)

        if waveform == 'Sine':
            wave = amp * np.sin(2 * np.pi * t + phi)
        elif waveform == 'Square':
            wave = amp * np.sign(np.sin(2 * np.pi * t + phi))
            wave[wave == 0] = amp  # ゼロクロス点を正側に丸める
        elif waveform == 'Triangle':
            # 鋸波をarcsinで三角波化
            wave = amp * (2 / np.pi) * np.arcsin(np.sin(2 * np.pi * t + phi))
        elif waveform == 'Sawtooth':
            # 位相をt軸にシフトして実現
            t_shift = (t + phase_deg / 360.0) % 1.0
            wave = amp * (2 * t_shift - 1)
        else:
            wave = np.zeros(n)

        wave = np.clip(wave + offset, -10.0, 10.0)
        return wave

    def toggle_ao_output(self, ch_idx):
        """出力ボタン押下で出力 ON/OFF をトグルする"""
        if self.ao_running[ch_idx]:
            self._stop_ao(ch_idx)
        else:
            self._start_ao(ch_idx)

    def _start_ao(self, ch_idx):
        """AO ch_idx の出力を開始する"""
        try:
            # 既存タスクがあれば停止
            self._stop_ao(ch_idx)

            wave = self._make_waveform(ch_idx)
            dev  = self.config['device']['device_name']

            task = nidaqmx.Task()
            task.ao_channels.add_ao_voltage_chan(
                f"{dev}/ao{ch_idx}",
                min_val=-10.0, max_val=10.0)

            waveform_type = self.ao_waveform_vars[ch_idx].get()
            if waveform_type == 'DC':
                # DC: 単一値を連続出力（タイミング設定不要）
                task.write(float(wave[0]))
                task.start()
            else:
                freq = max(0.01, self.ao_freq_vars[ch_idx].get())
                n = len(wave)
                # 1周期分のバッファを連続再生
                task.timing.cfg_samp_clk_timing(
                    rate=self.AO_RATE,
                    sample_mode=nidaqmx.constants.AcquisitionType.CONTINUOUS,
                    samps_per_chan=n)
                task.write(wave, auto_start=False)
                task.start()

            self.ao_tasks[ch_idx]   = task
            self.ao_running[ch_idx] = True

            # ボタンを「出力中」の見た目に変更
            btn = self.ao_out_btns[ch_idx]
            btn.config(text=f"⏹ Stop AO{ch_idx}",
                       bg='#cc3333', fg='white', relief=tk.SUNKEN)

        except Exception as e:
            messagebox.showerror("AO Error",
                                 f"Failed to start AO{ch_idx}:\n{e}")
            self._stop_ao(ch_idx)

    def _stop_ao(self, ch_idx):
        """AO ch_idx の出力を停止してタスクを解放する"""
        task = self.ao_tasks[ch_idx]
        if task is not None:
            try:
                task.stop()
                task.close()
            except Exception:
                pass
            self.ao_tasks[ch_idx] = None

        self.ao_running[ch_idx] = False

        # ボタンを「停止中」の見た目に戻す
        btn = self.ao_out_btns[ch_idx]
        if btn is not None:
            try:
                btn.config(text=f"▶ Output AO{ch_idx}",
                           bg='#d9d9d9', fg='black', relief=tk.RAISED)
            except Exception:
                pass

    # ================================================================== #
    # [New v2.4] Digital I/O 統合ロジック（P0.0〜P0.7）
    #   各ラインは 'in' / 'out' モード。モードに応じてDIタスク/DOタスクを
    #   動的に構築。DIの読取りは update_plot の描画タイミングで実施。
    # ================================================================== #
    def toggle_dio_mode(self, line):
        """ラインのモードを IN/OUT で切り替える。"""
        if self.dio_mode[line] == 'out':
            self.dio_mode[line] = 'in'
        else:
            self.dio_mode[line] = 'out'
            # OUTに戻したらOFFから開始
            self.dio_out_state[line] = False
        self._update_dio_mode_button(line)
        self._update_dio_state_button(line)
        self._rebuild_dio_tasks()

    def on_dio_state_click(self, line):
        """状態ボタンのクリック。OUTラインのみトグル出力する（INは無反応）。"""
        if self.dio_mode[line] != 'out':
            return
        self.dio_out_state[line] = not self.dio_out_state[line]
        self._write_dio_out()
        self._update_dio_state_button(line)

    def _update_dio_mode_button(self, line):
        """モードボタンの見た目を更新（IN=水色, OUT=橙）。"""
        mb = self.dio_mode_btns[line]
        if mb is None:
            return
        if self.dio_mode[line] == 'in':
            mb.config(text="IN", bg='#2aa6c4', fg='white')
        else:
            mb.config(text="OUT", bg='#e8923a', fg='white')

    def _update_dio_state_button(self, line):
        """状態ボタンの見た目を更新（OFF/LOW=黄, ON/HIGH=緑）。"""
        sb = self.dio_state_btns[line]
        if sb is None:
            return
        if self.dio_mode[line] == 'out':
            on = self.dio_out_state[line]
            relief = tk.SUNKEN if on else tk.RAISED
        else:
            on = self.dio_in_state[line]
            relief = tk.FLAT
        if on:
            sb.config(text=f"{line}", bg='#33aa33', fg='white', relief=relief)
        else:
            sb.config(text=f"{line}", bg='#f0d000', fg='black', relief=relief)

    def _rebuild_dio_tasks(self):
        """現在のモード割り当てに基づき DO/DI タスクを作り直す。"""
        # 既存タスクを解放
        self._close_do_task()
        self._close_di_task()

        out_lines = [i for i in range(8) if self.dio_mode[i] == 'out']
        in_lines  = [i for i in range(8) if self.dio_mode[i] == 'in']
        dev = self.config['device']['device_name']

        # DOタスク
        self._do_lines = out_lines
        if out_lines:
            try:
                self.do_task = nidaqmx.Task()
                for ln in out_lines:
                    self.do_task.do_channels.add_do_chan(f"{dev}/port0/line{ln}")
                self.do_task.start()
                self._write_dio_out()
            except Exception as e:
                print(f"[DIO] DO build error: {e}")
                self._close_do_task()

        # DIタスク
        self._di_lines = in_lines
        if in_lines:
            try:
                self.di_task = nidaqmx.Task()
                for ln in in_lines:
                    self.di_task.di_channels.add_di_chan(f"{dev}/port0/line{ln}")
                self.di_task.start()
            except Exception as e:
                print(f"[DIO] DI build error: {e}")
                self._close_di_task()

    def _write_dio_out(self):
        """OUTラインの状態をハードウェアへ書き込む。"""
        if self.do_task is None or not self._do_lines:
            return
        try:
            vals = [self.dio_out_state[ln] for ln in self._do_lines]
            self.do_task.write(vals if len(vals) > 1 else vals[0])
        except Exception as e:
            print(f"[DIO] DO write error: {e}")

    def _read_dio_in(self):
        """INラインを1回読み取り、状態ボタンを更新する。
        update_plot の描画タイミングから呼ばれる（別タイマー不要・軽量）。
        """
        if self.di_task is None or not self._di_lines:
            return
        try:
            vals = self.di_task.read()
            if not isinstance(vals, list):
                vals = [vals]
            for idx, ln in enumerate(self._di_lines):
                state = bool(vals[idx])
                if state != self.dio_in_state[ln]:   # 変化時のみUI更新
                    self.dio_in_state[ln] = state
                    self._update_dio_state_button(ln)
        except Exception as e:
            print(f"[DIO] DI read error: {e}")

    def _close_do_task(self):
        if self.do_task is not None:
            try:
                if self._do_lines:
                    lows = [False] * len(self._do_lines)
                    self.do_task.write(lows if len(lows) > 1 else lows[0])
                self.do_task.stop(); self.do_task.close()
            except Exception:
                pass
            self.do_task = None

    def _close_di_task(self):
        if self.di_task is not None:
            try:
                self.di_task.stop(); self.di_task.close()
            except Exception:
                pass
            self.di_task = None

    def _stop_dio(self):
        """終了時：DO/DIタスクを両方解放する。"""
        self._close_do_task()
        self._close_di_task()

    # ================================================================== #
    # [New v2.1] Frequency Response (Bode) 測定ロジック
    # ================================================================== #
    def _parse_fr_source(self, label):
        """信号源ラベルを ('ao', 0) または ('ai', ch_idx) のタプルに変換。"""
        if label.startswith('AO0'):
            return ('ao', 0)
        elif label.startswith('AI'):
            ch = int(label[2])
            return ('ai', ch)
        return ('ao', 0)

    def toggle_freq_response(self):
        """測定の開始/中断をトグルする。"""
        if self.fr_running:
            self.fr_running = False   # 中断要求（測定ループが検知して停止）
            self.fr_status_label.config(text="Stopping...", foreground='orange')
        else:
            self.start_freq_response()

    def start_freq_response(self):
        """周波数特性測定を開始する。"""
        try:
            f_start = self.fr_fstart_var.get()
            f_stop  = self.fr_fstop_var.get()
            n_pts   = self.fr_points_var.get()
            amp     = self.fr_amp_var.get()
            settle  = self.fr_settle_var.get()
            n_cyc   = self.fr_cycles_var.get()
            sweep   = self.fr_logsweep_var.get()

            if f_start <= 0 or f_stop <= f_start:
                messagebox.showerror("Error", "f stop must be greater than f start (>0).")
                return
            if f_stop > 500:
                if not messagebox.askyesno(
                        "Warning",
                        f"f stop = {f_stop} Hz exceeds the practical limit (~500 Hz) "
                        "for USB-6002 AO (5 kS/s).\nContinue anyway?"):
                    return

            input_src  = self._parse_fr_source(self.fr_input_var.get())
            output_src = self._parse_fr_source(self.fr_output_var.get())

            # 応答(と入力)にAIを使う場合、そのchがアクティブで取得対象である必要がある
            needed_ai = [s[1] for s in (input_src, output_src) if s[0] == 'ai']
            for ch in needed_ai:
                if not self.active_channels[ch]:
                    messagebox.showerror(
                        "Error",
                        f"AI{ch} is selected but not active.\n"
                        "Please enable it in Active Channels first.")
                    return

            # 周波数リスト生成
            if sweep == 'Log':
                freqs = np.logspace(np.log10(f_start), np.log10(f_stop), n_pts)
            else:
                freqs = np.linspace(f_start, f_stop, n_pts)

            self.fr_params = {
                'freqs': freqs, 'amp': amp, 'settle': settle,
                'n_cyc': n_cyc, 'input_src': input_src,
                'output_src': output_src,
                'input_label': self.fr_input_var.get(),
                'output_label': self.fr_output_var.get(),
            }

            # メインの監視ループを一時停止（DAQリソース競合を避ける）
            self._fr_was_paused = self.is_paused
            self.fr_running = True
            self.fr_run_btn.config(text="⏹ Stop Measurement", bg='#cc3333')

            # 測定はバックグラウンドの after チェーンで進める（UIを固めない）
            self._fr_index = 0
            self._fr_gain = []
            self._fr_phase = []
            self._fr_done_freqs = []
            self.root.after(100, self._fr_measure_next)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to start measurement:\n{e}")
            self._fr_finish(aborted=True)

    def _fr_measure_next(self):
        """1周波数ぶんの測定を行い、次へ進む（after チェーン）。"""
        if not self.fr_running:
            self._fr_finish(aborted=True)
            return

        params = self.fr_params
        freqs  = params['freqs']
        if self._fr_index >= len(freqs):
            self._fr_finish(aborted=False)
            return

        freq = float(freqs[self._fr_index])
        self.fr_status_label.config(
            text=f"Measuring {self._fr_index + 1}/{len(freqs)}: {freq:.2f} Hz",
            foreground='blue')
        self.root.update_idletasks()

        try:
            gain_db, phase_deg = self._fr_measure_single(freq, params)
            self._fr_done_freqs.append(freq)
            self._fr_gain.append(gain_db)
            self._fr_phase.append(phase_deg)
        except Exception as e:
            print(f"[FreqResp] {freq:.2f}Hz error: {e}")

        self._fr_index += 1
        # 次の周波数へ（わずかな遅延でUIイベントを処理）
        self.root.after(30, self._fr_measure_next)

    def _fr_measure_single(self, freq, params):
        """単一周波数で加振→取得→正弦波フィット→ゲイン/位相を返す。"""
        AO_RATE = 5000
        amp     = params['amp']
        settle  = params['settle']
        n_cyc   = params['n_cyc']
        in_src  = params['input_src']
        out_src = params['output_src']

        # 取得サンプル数：測定サイクル分（最低でも数百点）
        meas_time = n_cyc / freq
        ai_rate   = self.config['device']['sampling_rate']
        n_meas    = max(int(meas_time * ai_rate), 64)

        # ── AO0 で連続正弦波を出力 ──────────────────────────── #
        n_ao = max(8, int(round(AO_RATE / freq)))
        t_ao = np.arange(n_ao) / AO_RATE
        wave = amp * np.sin(2 * np.pi * freq * t_ao)
        wave = np.clip(wave, -10.0, 10.0)

        dev = self.config['device']['device_name']

        # 既存のメインタスクを一時停止（AI競合回避）
        main_paused_here = False
        if self.task is not None:
            try:
                self.task.stop()
                main_paused_here = True
            except Exception:
                pass

        ao_task = nidaqmx.Task()
        ai_task = nidaqmx.Task()
        try:
            ao_task.ao_channels.add_ao_voltage_chan(
                f"{dev}/ao0", min_val=-10.0, max_val=10.0)
            ao_task.timing.cfg_samp_clk_timing(
                rate=AO_RATE,
                sample_mode=AcquisitionType.CONTINUOUS,
                samps_per_chan=n_ao)
            ao_task.write(wave, auto_start=False)
            ao_task.start()

            # セトリング待ち（過渡応答を捨てる）
            import time
            time.sleep(settle)

            # ── AI 取得 ────────────────────────────────────── #
            # 入力参照・応答に必要なAIチャンネルを取得
            ai_channels = []
            ai_map = {}   # ch_idx -> 取得配列インデックス
            for src in (in_src, out_src):
                if src[0] == 'ai' and src[1] not in ai_map:
                    ai_map[src[1]] = len(ai_channels)
                    ai_channels.append(src[1])

            t_meas = np.arange(n_meas) / ai_rate

            if ai_channels:
                for ch in ai_channels:
                    ai_task.ai_channels.add_ai_voltage_chan(
                        f"{dev}/ai{ch}", min_val=-10.0, max_val=10.0)
                ai_task.timing.cfg_samp_clk_timing(
                    rate=ai_rate,
                    sample_mode=AcquisitionType.FINITE,
                    samps_per_chan=n_meas)
                ai_task.start()
                ai_data = ai_task.read(
                    number_of_samples_per_channel=n_meas, timeout=10.0)
                ai_task.stop()
                if not isinstance(ai_data[0], list):
                    ai_data = [ai_data]
                ai_data = [np.asarray(d) for d in ai_data]

            # ── 入力参照信号を用意 ──────────────────────────── #
            if in_src[0] == 'ao':
                # AO0指令そのもの（理想正弦波）を参照に使う
                in_sig = amp * np.sin(2 * np.pi * freq * t_meas)
            else:
                idx = ai_map[in_src[1]]
                raw = ai_data[idx]
                in_sig = raw * self.config['channels'][in_src[1]]['conversion_factor']

            # ── 応答信号 ────────────────────────────────────── #
            if out_src[0] == 'ao':
                out_sig = amp * np.sin(2 * np.pi * freq * t_meas)
            else:
                idx = ai_map[out_src[1]]
                raw = ai_data[idx]
                out_sig = raw * self.config['channels'][out_src[1]]['conversion_factor']

        finally:
            for tk_task in (ao_task, ai_task):
                try:
                    tk_task.stop(); tk_task.close()
                except Exception:
                    pass
            # メインタスクを再開
            if main_paused_here and self.task is not None:
                try:
                    self.task.start()
                except Exception:
                    pass

        # ── 単一周波数正弦波フィット（最小二乗）───────────────── #
        a_in,  b_in,  _ = self._fit_sine(t_meas, in_sig, freq)
        a_out, b_out, _ = self._fit_sine(t_meas, out_sig, freq)

        amp_in  = np.hypot(a_in, b_in)
        amp_out = np.hypot(a_out, b_out)
        ph_in   = np.arctan2(b_in, a_in)
        ph_out  = np.arctan2(b_out, a_out)

        if amp_in < 1e-9:
            gain_db = float('nan')
        else:
            gain_db = 20 * np.log10(amp_out / amp_in)

        phase_deg = np.rad2deg(ph_out - ph_in)
        # -180〜180 に正規化
        phase_deg = (phase_deg + 180) % 360 - 180

        return gain_db, phase_deg

    @staticmethod
    def _fit_sine(t, y, freq):
        """y ≈ a·cos(2πft) + b·sin(2πft) + c を最小二乗で解く。"""
        w = 2 * np.pi * freq
        A = np.column_stack([np.cos(w * t), np.sin(w * t), np.ones_like(t)])
        coef, *_ = np.linalg.lstsq(A, y, rcond=None)
        return coef[0], coef[1], coef[2]   # a, b, c

    def _fr_finish(self, aborted=False):
        """測定終了処理：結果保存とボード線図表示。"""
        self.fr_running = False
        self.fr_run_btn.config(text="▶ Measure Frequency Response", bg='#3366cc')

        if aborted or len(self._fr_done_freqs) == 0:
            self.fr_status_label.config(text="Aborted / no data", foreground='gray')
            return

        self.fr_results = {
            'freq':  np.asarray(self._fr_done_freqs),
            'gain':  np.asarray(self._fr_gain),
            'phase': np.asarray(self._fr_phase),
            'input_label':  self.fr_params['input_label'],
            'output_label': self.fr_params['output_label'],
        }
        self.fr_status_label.config(
            text=f"Done: {len(self._fr_done_freqs)} points",
            foreground='green')

        # Excel保存 → ボード線図表示
        self._save_freq_response_excel()
        self.show_bode_window()

    def _save_freq_response_excel(self):
        """測定結果をExcelに保存（データ＋ボード線図画像）。"""
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            r = self.fr_results
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BodeData"

            hdr_fill = PatternFill("solid", fgColor="1F4E79")
            hdr_font = Font(bold=True, color="FFFFFF", size=11)
            thin = Side(style='thin', color='AAAAAA')

            headers = ["Frequency[Hz]", "Gain[dB]", "Phase[deg]"]
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(c)].width = 16

            for i in range(len(r['freq'])):
                ws.cell(row=i + 2, column=1, value=round(float(r['freq'][i]), 4))
                ws.cell(row=i + 2, column=2, value=round(float(r['gain'][i]), 4))
                ws.cell(row=i + 2, column=3, value=round(float(r['phase'][i]), 4))

            # メタ情報
            ws.cell(row=1, column=5, value="Input ref:")
            ws.cell(row=1, column=6, value=r['input_label'])
            ws.cell(row=2, column=5, value="Response:")
            ws.cell(row=2, column=6, value=r['output_label'])

            # ボード線図画像
            img_buf = self._render_bode_png()
            if img_buf is not None:
                ws_g = wb.create_sheet("BodePlot")
                xl_img = XLImage(img_buf)
                xl_img.width = 720
                xl_img.height = 560
                ws_g.add_image(xl_img, "B2")

            save_dir = r"C:\PRG\data"
            os.makedirs(save_dir, exist_ok=True)
            now_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            save_path = os.path.join(save_dir, f"FreqResponse_{now_str}.xlsx")
            wb.save(save_path)
            print(f"[FreqResp] Saved: {save_path}")
        except Exception as e:
            print(f"[FreqResp] Excel save error: {e}")
            import traceback; traceback.print_exc()

    def _render_bode_png(self):
        """現在の結果とフォーマット設定でボード線図PNGを生成（BytesIO）。"""
        if self.fr_results is None:
            return None
        try:
            r   = self.fr_results
            cfg = self.fr_graph_cfg
            ff  = cfg['font_family']

            fig, (ax_g, ax_p) = plt.subplots(
                2, 1, figsize=(8, 6.2), sharex=True)
            fig.set_facecolor('white')

            ax_g.semilogx(r['freq'], r['gain'],
                          'o-', color='#1f6fc4',
                          linewidth=cfg['line_width'],
                          markersize=cfg['marker_size'])
            ax_g.set_ylabel("Gain [dB]", fontsize=cfg['font_size_label'],
                            fontweight='bold', fontfamily=ff)
            ax_g.set_title(cfg['title'], fontsize=cfg['font_size_title'],
                           fontweight='bold', fontfamily=ff)

            ax_p.semilogx(r['freq'], r['phase'],
                          's-', color='#c43f1f',
                          linewidth=cfg['line_width'],
                          markersize=cfg['marker_size'])
            ax_p.set_ylabel("Phase [deg]", fontsize=cfg['font_size_label'],
                            fontweight='bold', fontfamily=ff)
            ax_p.set_xlabel("Frequency [Hz]", fontsize=cfg['font_size_label'],
                            fontweight='bold', fontfamily=ff)

            for ax in (ax_g, ax_p):
                ax.tick_params(labelsize=cfg['font_size_tick'])
                if cfg['grid']:
                    ax.grid(True, which='both', alpha=0.3)
                for sp in ax.spines.values():
                    sp.set_linewidth(1.2)

            fig.tight_layout()
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=150,
                        facecolor='white', bbox_inches='tight')
            buf.seek(0)
            plt.close(fig)
            return buf
        except Exception as e:
            print(f"[FreqResp] Bode render error: {e}")
            return None

    def show_bode_window(self):
        """ボード線図を別ウィンドウで表示（フォーマット変更UI付き）。"""
        if self.fr_results is None:
            messagebox.showinfo("Info", "No measurement data yet.")
            return

        win = tk.Toplevel(self.root)
        win.title("Bode Diagram")
        win.geometry("900x720")

        # 上部：フォーマット変更パネル
        fmt = ttk.LabelFrame(win, text="Graph Format", padding="6")
        fmt.pack(fill=tk.X, padx=8, pady=6)

        cfg = self.fr_graph_cfg
        title_var = tk.StringVar(value=cfg['title'])
        font_var  = tk.StringVar(value=cfg['font_family'])
        tsize_var = tk.IntVar(value=cfg['font_size_title'])
        lsize_var = tk.IntVar(value=cfg['font_size_label'])
        ksize_var = tk.IntVar(value=cfg['font_size_tick'])
        lw_var    = tk.DoubleVar(value=cfg['line_width'])
        ms_var    = tk.DoubleVar(value=cfg['marker_size'])
        grid_var  = tk.BooleanVar(value=cfg['grid'])

        row1 = ttk.Frame(fmt); row1.pack(fill=tk.X, pady=2)
        ttk.Label(row1, text="Title:").pack(side=tk.LEFT)
        ttk.Entry(row1, textvariable=title_var, width=22).pack(side=tk.LEFT, padx=4)
        ttk.Label(row1, text="Font:").pack(side=tk.LEFT)
        ttk.Combobox(row1, textvariable=font_var, width=12, state='readonly',
                     values=['Arial', 'Times New Roman', 'Calibri',
                             'DejaVu Sans', 'MS Gothic']).pack(side=tk.LEFT, padx=4)
        ttk.Checkbutton(row1, text="Grid", variable=grid_var).pack(side=tk.LEFT, padx=4)

        row2 = ttk.Frame(fmt); row2.pack(fill=tk.X, pady=2)
        for lbl, var, w in [("Title size", tsize_var, 5),
                            ("Label size", lsize_var, 5),
                            ("Tick size", ksize_var, 5),
                            ("Line W", lw_var, 5),
                            ("Marker", ms_var, 5)]:
            ttk.Label(row2, text=lbl + ":").pack(side=tk.LEFT)
            ttk.Spinbox(row2, textvariable=var, from_=0, to=40,
                        increment=1, width=w).pack(side=tk.LEFT, padx=(2, 8))

        # グラフ表示領域
        graph_holder = ttk.Frame(win)
        graph_holder.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        canvas_holder = {'canvas': None, 'fig': None}

        def redraw():
            # 設定を反映
            cfg['title']           = title_var.get()
            cfg['font_family']     = font_var.get()
            cfg['font_size_title'] = tsize_var.get()
            cfg['font_size_label'] = lsize_var.get()
            cfg['font_size_tick']  = ksize_var.get()
            cfg['line_width']      = lw_var.get()
            cfg['marker_size']     = ms_var.get()
            cfg['grid']            = grid_var.get()

            if canvas_holder['canvas'] is not None:
                canvas_holder['canvas'].get_tk_widget().destroy()
                plt.close(canvas_holder['fig'])

            r = self.fr_results
            ff = cfg['font_family']
            fig, (ax_g, ax_p) = plt.subplots(2, 1, figsize=(8, 5.6), sharex=True)
            fig.set_facecolor('white')
            ax_g.semilogx(r['freq'], r['gain'], 'o-', color='#1f6fc4',
                          linewidth=cfg['line_width'], markersize=cfg['marker_size'])
            ax_g.set_ylabel("Gain [dB]", fontsize=cfg['font_size_label'],
                            fontweight='bold', fontfamily=ff)
            ax_g.set_title(cfg['title'], fontsize=cfg['font_size_title'],
                           fontweight='bold', fontfamily=ff)
            ax_p.semilogx(r['freq'], r['phase'], 's-', color='#c43f1f',
                          linewidth=cfg['line_width'], markersize=cfg['marker_size'])
            ax_p.set_ylabel("Phase [deg]", fontsize=cfg['font_size_label'],
                            fontweight='bold', fontfamily=ff)
            ax_p.set_xlabel("Frequency [Hz]", fontsize=cfg['font_size_label'],
                            fontweight='bold', fontfamily=ff)
            for ax in (ax_g, ax_p):
                ax.tick_params(labelsize=cfg['font_size_tick'])
                if cfg['grid']:
                    ax.grid(True, which='both', alpha=0.3)
            fig.tight_layout()

            cv = FigureCanvasTkAgg(fig, master=graph_holder)
            cv.draw()
            cv.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            canvas_holder['canvas'] = cv
            canvas_holder['fig'] = fig

        # ボタン列
        btns = ttk.Frame(fmt); btns.pack(fill=tk.X, pady=(4, 0))
        ttk.Button(btns, text="🔄 Apply Format", command=redraw).pack(side=tk.LEFT)
        ttk.Button(btns, text="💾 Save PNG...",
                   command=lambda: self._save_bode_figure()).pack(side=tk.LEFT, padx=6)

        redraw()

    def _save_bode_figure(self):
        """ボード線図を画像ファイルに保存する（拡張子に応じた形式）。"""
        if self.fr_results is None:
            return
        filename = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG Image", "*.png"), ("PDF Vector", "*.pdf"),
                       ("SVG Vector", "*.svg")],
            initialfile=f"Bode_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}")
        if not filename:
            return
        try:
            r   = self.fr_results
            cfg = self.fr_graph_cfg
            ff  = cfg['font_family']
            fig, (ax_g, ax_p) = plt.subplots(2, 1, figsize=(8, 6.2), sharex=True)
            fig.set_facecolor('white')
            ax_g.semilogx(r['freq'], r['gain'], 'o-', color='#1f6fc4',
                          linewidth=cfg['line_width'], markersize=cfg['marker_size'])
            ax_g.set_ylabel("Gain [dB]", fontsize=cfg['font_size_label'],
                            fontweight='bold', fontfamily=ff)
            ax_g.set_title(cfg['title'], fontsize=cfg['font_size_title'],
                           fontweight='bold', fontfamily=ff)
            ax_p.semilogx(r['freq'], r['phase'], 's-', color='#c43f1f',
                          linewidth=cfg['line_width'], markersize=cfg['marker_size'])
            ax_p.set_ylabel("Phase [deg]", fontsize=cfg['font_size_label'],
                            fontweight='bold', fontfamily=ff)
            ax_p.set_xlabel("Frequency [Hz]", fontsize=cfg['font_size_label'],
                            fontweight='bold', fontfamily=ff)
            for ax in (ax_g, ax_p):
                ax.tick_params(labelsize=cfg['font_size_tick'])
                if cfg['grid']:
                    ax.grid(True, which='both', alpha=0.3)
            fig.tight_layout()
            dpi = 300 if filename.lower().endswith(('.png',)) else None
            fig.savefig(filename, dpi=dpi, facecolor='white', bbox_inches='tight')
            plt.close(fig)
            messagebox.showinfo("Success", f"Bode diagram saved!\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save:\n{e}")

    def on_closing(self):
        # ------------------------------------------------------------------ #
        # [Fix 1] active_channels を config に書き込んでから保存
        # ------------------------------------------------------------------ #
        self.config['active_channels'] = [
            var.get() for var in self.channel_vars]

        # チャンネル名・変換係数（メインパネルで直接編集された値）も反映
        for i in range(8):
            try:
                self.config['channels'][i]['name'] = \
                    self.channel_name_vars[i].get()
                self.config['channels'][i]['conversion_factor'] = \
                    self.channel_factor_vars[i].get()
            except Exception:
                pass

        try:
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Warning: Failed to save config: {e}")

        self.cleanup()
        self.root.quit()
        self.root.destroy()


# ========================================================================== #
def main():
    root = tk.Tk()
    app = DAQApplication(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()


if __name__ == "__main__":
    main()
